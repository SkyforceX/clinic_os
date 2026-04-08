from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.contract.models import ImplementationPlan


THIN = Side(style="thin", color="C9D2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="EAF3FB")


IMPLEMENTATION_TEMPLATE = [
    {
        "owner": "Các BP",
        "category": "__COMPANY_NAME__",
        "detail_type": "company_short_name",
        "detail_default": "",
        "note_default": "Để các BP nhận biết khi giao tiếp với KH",
    },
    {
        "owner": "Các BP",
        "category": "ĐỊA CHỈ",
        "detail_type": "company_address",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Kế toán",
        "category": "MST",
        "detail_type": "tax_code",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Các BP",
        "category": "NGƯỜI LIÊN HỆ",
        "detail_type": "contact_person",
        "detail_default": "",
        "note_default": "Liên hệ khi có phát sinh sự việc",
    },
    {
        "owner": "Các BP",
        "category": "SỐ LƯỢNG NHÂN VIÊN (LẤY MÁU)",
        "detail_type": "employee_count",
        "detail_default": "",
        "note_default": "Chi tiết theo Danh sách đính kèm",
    },
    {
        "owner": "VH + ĐD + Sales",
        "category": "THỜI GIAN LẤY MÁU",
        "detail_type": "blood_collection_time",
        "detail_default": "",
        "note_default": "Các ca chưa lấy máu tại VP CTY, sẽ ghé VMD lấy",
    },
    {
        "owner": "VH + ĐD + Sales",
        "category": "ĐỊA ĐIỂM LẤY MÁU",
        "detail_type": "blood_collection_location",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Các BP",
        "category": "THỜI GIAN KHÁM",
        "detail_type": "checkup_time",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "DANH MỤC KHÁM NGOÀI GÓI",
        "detail_type": "fixed",
        "detail_default": "TOÀN BỘ CHI PHÍ PHÁT SINH của Nhân viên và Người thân sẽ tự thanh toán.",
        "note_default": "Người thân sẽ do CTY thông báo",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "ƯU ĐÃI: NGƯỜI THÂN",
        "detail_type": "fixed",
        "detail_default": "Áp dụng cho người thân theo Danh sách Cty đã xác nhận",
        "note_default": "TỰ THANH TOÁN",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "ƯU ĐÃI GIẢM GIÁ (áp dụng cho nhân viên và Người thân SAU KHI đã sử dụng Gói khám CTY)",
        "detail_type": "fixed",
        "detail_default": "Giảm 10% các dịch vụ ngoài Gói CTY, ngoại trừ (Thuốc, Nội Soi, Phẫu thuật, Thủ Thuật, các dịch vụ có sử dụng Vật tư tiêu hao đặc biệt.)",
        "note_default": "TỰ THANH TOÁN",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "PHÁT SINH CHUYỂN ĐỔI GÓI: NỮ ĐT --> <-- NỮ GĐ",
        "detail_type": "fixed",
        "detail_default": "Nhân viên ĐƯỢC / KHÔNG chuyển đổi Gói Nữ ĐT sang Gói Nữ GĐ",
        "note_default": "",
    },
    {
        "owner": "Sales - TKYK",
        "category": "CTY CÓ YÊU CẦU CUNG CẤP GIẤY CNSK TT32 ?",
        "detail_type": "fixed",
        "detail_default": "KHÔNG YÊU CẦU / CÓ YÊU CẦU",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "THỜI HẠN & HÌNH THỨC TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "Thời hạn: 10 ngày, kể từ ngày KH đã hoàn tất Gói khám\nHình thức: TRẢ 1 LẦN / CUỐN CHIẾU.",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "ĐỊA CHỈ TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "* Hồ sơ cá nhân:\n* SMR:",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "NGƯỜI LIÊN HỆ TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "Sales sẽ thông báo",
        "note_default": "",
    },
    {
        "owner": "BS - Điều Dưỡng",
        "category": "PHÂN LOẠI KHÁM SỨC KHỎE",
        "detail_type": "fixed",
        "detail_default": "Có",
        "note_default": "",
    },
    {
        "owner": "TKYK - BS - Điều Dưỡng",
        "category": "BÁO CÁO TỔNG KẾT (SUMMARY REPORT)",
        "detail_type": "fixed",
        "detail_default": "Tiếng Việt",
        "note_default": "",
    },
    {
        "owner": "BS - Điều Dưỡng",
        "category": "PHÁT HIỆN BỆNH LÝ",
        "detail_type": "fixed",
        "detail_default": "BS / Điều dưỡng thông tin cho Bệnh nhân và Sales (nếu có liên quan đến Bệnh truyền nhiễm).",
        "note_default": "",
    },
    {
        "owner": "Kế toán - Sales",
        "category": "THỜI HẠN THANH TOÁN",
        "detail_type": "fixed",
        "detail_default": "10 ngày, kể từ ngày Kế toán xuất hóa đơn tài chính",
        "note_default": "",
    },
    {
        "owner": "CSKH - Sales",
        "category": "BÁO CÁO ĐÁNH GIÁ DỊCH VỤ TỪ KHÁCH HÀNG",
        "detail_type": "fixed",
        "detail_default": "Gửi KH Form Đánh giá dịch vụ --> Ghi nhận ý kiến KH --> QL Chất lượng sau khi hoàn tất KSK",
        "note_default": "",
    },
]


def _clean_spaces(value):
    return re.sub(r"\s+", " ", (value or "").strip())


def build_company_short_name(name: str) -> str:
    raw = _clean_spaces(name).upper()
    if not raw:
        return ""

    stop_words = {
        "CÔNG",
        "TY",
        "TNHH",
        "MTV",
        "TM",
        "DV",
        "DỊCH",
        "VỤ",
        "CP",
        "CỔ",
        "PHẦN",
        "DOANH",
        "NGHIỆP",
        "TRÁCH",
        "NHIỆM",
        "HỮU",
        "HẠN",
        "CHI",
        "NHÁNH",
        "VÀ",
    }
    words = [w for w in re.split(r"[^\wÀ-ỹ]+", raw) if w and w not in stop_words]
    if not words:
        return raw[:30]

    short_name = "".join(word[0] for word in words[:6])
    return short_name[:20]


def build_contact_display(contract) -> str:
    user = getattr(contract, "created_by", None)
    if not user:
        return ""
    return getattr(user, "get_full_name", lambda: "")() or getattr(user, "username", "") or ""


def format_date(value):
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def format_datetime(value):
    if not value:
        return ""
    return value.strftime("%H:%M, %d/%m/%Y")


def build_blood_collection_time(contract, profile):
    if profile and getattr(profile, "blood_collection_from_at", None):
        if getattr(profile, "blood_collection_to_at", None):
            return f"{format_datetime(profile.blood_collection_from_at)} - {format_datetime(profile.blood_collection_to_at)}"
        return format_datetime(profile.blood_collection_from_at)

    first_row = contract.blood_collection_schedules.order_by("collection_date", "id").first()
    if first_row and first_row.collection_date:
        return format_date(first_row.collection_date)
    return ""


def build_blood_collection_location(contract, profile):
    if profile and getattr(profile, "blood_collection_location", None):
        return profile.blood_collection_location or ""
    first_row = contract.blood_collection_schedules.order_by("collection_date", "id").first()
    return getattr(first_row, "location", "") or ""


def build_checkup_time(contract):
    if contract.start_date and contract.end_date:
        return f"{format_date(contract.start_date)} - {format_date(contract.end_date)}"
    if contract.start_date:
        return format_date(contract.start_date)
    if contract.end_date:
        return format_date(contract.end_date)
    return ""


def build_employee_count(contract, profile):
    if contract.employee_count:
        return str(contract.employee_count)

    if profile:
        total = (
            int(getattr(profile, "male_count", 0) or 0)
            + int(getattr(profile, "female_single_count", 0) or 0)
            + int(getattr(profile, "female_family_count", 0) or 0)
        )
        if total > 0:
            return str(total)

    return ""


def build_dynamic_value(contract, profile, detail_type, category):
    company = contract.company
    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif company:
        company_name = getattr(company, "name", "") or ""

    if detail_type == "company_short_name":
        short_name = build_company_short_name(company_name)
        return f"Gọi tắt: {short_name}" if short_name else ""
    if detail_type == "company_address":
        if profile and getattr(profile, "company_address_snapshot", None):
            return profile.company_address_snapshot or ""
        return getattr(company, "address", "") or ""
    if detail_type == "tax_code":
        if profile and getattr(profile, "company_tax_code_snapshot", None):
            return profile.company_tax_code_snapshot or ""
        return getattr(company, "tax_code", "") or ""
    if detail_type == "contact_person":
        return build_contact_display(contract)
    if detail_type == "employee_count":
        return build_employee_count(contract, profile)
    if detail_type == "blood_collection_time":
        return build_blood_collection_time(contract, profile)
    if detail_type == "blood_collection_location":
        return build_blood_collection_location(contract, profile)
    if detail_type == "checkup_time":
        return build_checkup_time(contract)
    return ""


def build_default_rows(contract):
    profile = getattr(contract, "corporate_profile", None)
    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif contract.company:
        company_name = contract.company.name

    rows = []
    for idx, item in enumerate(IMPLEMENTATION_TEMPLATE, start=1):
        category = item["category"]
        if category == "__COMPANY_NAME__":
            category = company_name or "TÊN CÔNG TY"

        detail = item["detail_default"]
        if item["detail_type"] != "fixed":
            detail = build_dynamic_value(contract, profile, item["detail_type"], category)

        rows.append(
            {
                "stt": idx,
                "owner": item["owner"],
                "category": category,
                "detail": detail or "",
                "note": item["note_default"] or "",
            }
        )
    return rows


def normalize_rows(rows, contract):
    defaults = build_default_rows(contract)
    if not rows:
        return defaults

    normalized = []
    for idx, default in enumerate(defaults):
        existing = rows[idx] if idx < len(rows) else {}
        normalized.append(
            {
                "stt": idx + 1,
                "owner": default["owner"],
                "category": default["category"],
                "detail": str(existing.get("detail", default["detail"]) or ""),
                "note": str(existing.get("note", default["note"]) or ""),
            }
        )
    return normalized


def get_or_create_plan(contract):
    plan, _ = ImplementationPlan.objects.get_or_create(contract=contract)
    normalized = normalize_rows(plan.rows_json or [], contract)
    if plan.rows_json != normalized:
        plan.rows_json = normalized
        plan.save(update_fields=["rows_json", "updated_at"])
    return plan


def update_plan_rows_from_post(plan, post_data, contract):
    defaults = build_default_rows(contract)

    details = post_data.getlist("row_detail[]")
    notes = post_data.getlist("row_note[]")

    rows = []
    for idx, default in enumerate(defaults):
        detail = details[idx].strip() if idx < len(details) and details[idx] is not None else default["detail"]
        note = notes[idx].strip() if idx < len(notes) and notes[idx] is not None else default["note"]

        rows.append(
            {
                "stt": idx + 1,
                "owner": default["owner"],
                "category": default["category"],
                "detail": detail,
                "note": note,
            }
        )

    plan.rows_json = rows
    plan.save(update_fields=["rows_json", "updated_at"])
    return plan


def build_package_rows(contract):
    rows = []
    details = contract.service_lines.order_by("display_order", "id")
    for idx, item in enumerate(details, start=1):
        rows.append(
            {
                "stt": idx,
                "group_name": item.group_name or "Khác",
                "item_name": item.item_name or "",
                "description": item.description or "",
                "male": bool(item.for_male),
                "female_single": bool(item.for_female_single),
                "female_family": bool(item.for_female_family),
            }
        )
    return rows


def build_sheet_year(contract):
    if contract.start_date:
        return contract.start_date.year
    if getattr(contract, "created_at", None):
        return contract.created_at.year
    return 2026


def _slugify_ascii(value):
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "ke-hoach-trien-khai"


def _apply_table_border(ws, start_row, end_row, start_col, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def export_plan_excel(contract, plan):
    year = build_sheet_year(contract)
    profile = getattr(contract, "corporate_profile", None)

    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif contract.company:
        company_name = contract.company.name

    wb = Workbook()
    ws1 = wb.active
    ws1.title = f"TRIỂN KHAI {year}"
    ws2 = wb.create_sheet(title=f"GÓI KHÁM {year}")

    rows = normalize_rows(plan.rows_json or [], contract)
    package_rows = build_package_rows(contract)

    # Sheet 1
    ws1.merge_cells("B1:E1")
    ws1["B1"] = f"TRIỂN KHAI KHÁM SỨC KHỎE NĂM {year}"
    ws1["B1"].font = Font(size=14, bold=True)
    ws1["B1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("B2:E2")
    short_name = build_company_short_name(company_name)
    line2 = company_name or ""
    if short_name:
        line2 += f"\n(VIẾT TẮT: {short_name})"
    ws1["B2"] = line2
    ws1["B2"].font = Font(size=12, bold=True)
    ws1["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1["A3"] = "NGƯỜI PHỤ TRÁCH"
    ws1["C3"] = build_contact_display(contract)
    ws1["A3"].font = Font(bold=True)
    ws1["C3"].font = Font(bold=True)

    ws1["A4"] = "STT"
    ws1["B4"] = "PHỤ TRÁCH"
    ws1["C4"] = "DANH MỤC"
    ws1["D4"] = "CHI TIẾT"
    ws1["E4"] = "GHI CHÚ / GIẢI THÍCH"
    _style_header_row(ws1, 4)

    current_row = 5
    for item in rows:
        ws1.cell(current_row, 1, item["stt"])
        ws1.cell(current_row, 2, item["owner"])
        ws1.cell(current_row, 3, item["category"])
        ws1.cell(current_row, 4, item["detail"])
        ws1.cell(current_row, 5, item["note"])
        current_row += 1

    _apply_table_border(ws1, 4, current_row - 1, 1, 5)

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 34
    ws1.column_dimensions["D"].width = 52
    ws1.column_dimensions["E"].width = 34

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 38

    for row_idx in range(5, current_row):
        ws1.row_dimensions[row_idx].height = 34

    # Sheet 2
    ws2.merge_cells("B1:F1")
    ws2["B1"] = f"DANH MỤC ĐĂNG KÝ KHÁM NĂM {year}\n{company_name}\n(VIẾT TẮT: {short_name})"
    ws2["B1"].font = Font(size=13, bold=True)
    ws2["B1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws2.merge_cells("B2:F2")
    ws2["B2"] = f"PHỤ TRÁCH: {build_contact_display(contract).upper()}"
    ws2["B2"].font = Font(bold=True)

    ws2.merge_cells("B3:F3")
    ws2["B3"] = f"SỐ LƯỢNG: {build_employee_count(contract, profile)} NGƯỜI"
    ws2["B3"].font = Font(bold=True)

    ws2.merge_cells("B4:F4")
    ws2["B4"] = f"THỜI GIAN: {build_checkup_time(contract)}"
    ws2["B4"].font = Font(bold=True)

    ws2["B5"] = "STT"
    ws2["C5"] = "KHÁM TỔNG QUÁT / GENERAL EXAMINATION"
    ws2["D5"] = "NAM"
    ws2["E5"] = "NỮ ĐỘC THÂN"
    ws2["F5"] = "NỮ CÓ GIA ĐÌNH"
    _style_header_row(ws2, 5)

    current_row = 6
    last_group = None
    for item in package_rows:
        if item["group_name"] != last_group:
            ws2.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
            cell = ws2.cell(current_row, 2, item["group_name"])
            cell.fill = GROUP_FILL
            cell.font = Font(bold=True, color="1F1F1F")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for c in range(2, 7):
                ws2.cell(current_row, c).border = BORDER
            current_row += 1
            last_group = item["group_name"]

        ws2.cell(current_row, 2, item["stt"])
        ws2.cell(current_row, 3, item["item_name"])
        ws2.cell(current_row, 4, "X" if item["male"] else "")
        ws2.cell(current_row, 5, "X" if item["female_single"] else "")
        ws2.cell(current_row, 6, "X" if item["female_family"] else "")
        for c in range(2, 7):
            ws2.cell(current_row, c).border = BORDER
            ws2.cell(current_row, c).alignment = Alignment(
                horizontal="center" if c in (2, 4, 5, 6) else "left",
                vertical="center",
                wrap_text=True,
            )
        current_row += 1

        if item["description"]:
            ws2.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
            ws2.cell(current_row, 3, item["description"])
            for c in range(2, 7):
                ws2.cell(current_row, c).border = BORDER
                ws2.cell(current_row, c).alignment = Alignment(vertical="center", wrap_text=True)
            current_row += 1

    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 66
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 16

    ws2.row_dimensions[1].height = 56

    for ws in (ws1, ws2):
        for row in ws.iter_rows():
            for cell in row:
                if cell.row in (1, 2, 3, 4, 5) and cell.value:
                    continue
                if cell.value is not None and not cell.border.left.style:
                    cell.border = BORDER
                if cell.value is not None and cell.alignment == Alignment():
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A4" if ws == ws1 else "B5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ke-hoach-trien-khai-{contract.contract_number or contract.pk}-{_slugify_ascii(company_name)}.xlsx"
    return output, filename