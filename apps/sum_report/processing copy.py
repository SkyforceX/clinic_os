import os
from pathlib import Path
from django.conf import settings
from django.http import FileResponse
from django.utils.text import slugify
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import re
import pickle
from pyvi import ViTokenizer, ViUtils
import io

# def demo():
#     # Giả sử bạn xác định dòng tiêu đề là dòng 1 (index=0), thì dùng skiprows=1
#     df_ksk = pd.read_excel("D:/Code/Project/Django project/VMD/unicore/media/nhathuduc/customer_list.xlsx", sheet_name="Danh sách", skiprows=1, dtype=str)
#     df_ksk.columns = df_ksk.columns.str.strip().str.lower()
#     # In tên cột thực tế
#     print("Tên cột thực tế:", df_ksk.columns.tolist())
# demo()

def kiem_tra_chinh_ta_viet(text):
    # Tách từ
    tokens = ViTokenizer.tokenize(text).split()
    # Danh sách các từ không hợp lệ (ví dụ: chứa ký tự lạ, không trong từ điển unicode tiếng Việt)
    loi = []
    for token in tokens:
        word = token.replace("_", " ")
        # Kiểm tra xem có chứa ký tự không hợp lệ không
        if not re.match(r"^[a-zA-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠàáâãèéêìíòóôõùúăđĩũơƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼỀỀỂưăạảấầẩẫậắằẳẵặẹẻẽềềểếỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪễệỉịọỏốồổỗộớờởỡợụủứừỬỮỰỲỴÝỶỸỳỵỷỹ\s\-]+$", word):
            loi.append(word)
    return loi


def save_df(df, filename):
    temp_dir = Path(settings.MEDIA_ROOT) / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)  # Tạo thư mục nếu chưa có

    output_path = temp_dir / filename  # Ghép đường dẫn tới file trong thư mục temp

    with open(output_path, "wb") as f:
        pickle.dump(df, f)

    return output_path  # Trả về đường dẫn để sử dụng tiếp nếu cần


def run_py1(test_list_file, out_pickle):
    df_raw = pd.read_excel(test_list_file, header=None, skiprows=16)

    # Tìm dòng dừng khi cột E chứa 'NGƯỜI LẬP BIỂU'
    stop_index = df_raw[df_raw[4].astype(str).str.contains("NGƯỜI LẬP BIỂU", na=False)].index
    if not stop_index.empty:
        df_raw = df_raw.loc[:stop_index[0] - 1]

    # ======= Xử lý dữ liệu =======
    data = []
    for _, row in df_raw.iterrows():
        hoten = row[3]
        yeucau = row[15]
        ketqua = row[16]

        tuoi_nam = row[5]
        tuoi_nu = row[7]

        gioitinh = None
        tuoi = None

        if pd.notna(tuoi_nam) and isinstance(tuoi_nam, (int, float)):
            gioitinh = "Nam"
            tuoi = tuoi_nam
        elif pd.notna(tuoi_nu) and isinstance(tuoi_nu, (int, float)):
            gioitinh = "Nữ"
            tuoi = tuoi_nu

        namsinh = int(datetime.now().year - tuoi) if tuoi else None

        data.append({
            'Họ tên người bệnh': hoten,
            'Giới tính': gioitinh,
            'Năm sinh': namsinh,
            'Yêu cầu': yeucau,
            'Kết quả': ketqua
        })

    # ======= Tạo DataFrame và ghi ra file Excel =======
    df = pd.DataFrame(data)
    # Lọc dữ liệu hợp lệ
    df = df[df['Họ tên người bệnh'].apply(lambda x: isinstance(x, str) and x.strip() != '')]
    df = df[df['Giới tính'].isin(['Nam', 'Nữ'])]
    df = df[df['Năm sinh'].notna()]

    # Lưu ra file pickle
    # save_df(df, "test_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df, out_pickle)


def run_py2(clinical_list_file, out_pickle):
    # ======== Đọc và xử lý dữ liệu từ Excel ===========    
    df = pd.read_excel(clinical_list_file)

    # Hàm xác định giới tính
    def get_gioitinh(row):
        if pd.notna(row[5]):
            return 'Nam'
        elif pd.notna(row[7]):
            return 'Nữ'
        return None

    # Hàm lấy năm sinh tại đúng cột giới tính
    def get_namsinh(row):
        if pd.notna(row[5]):
            return row[5]
        elif pd.notna(row[7]):
            return row[7]
        return None

    # Tạo DataFrame kết quả
    df_kq = pd.DataFrame()
    df_kq['Họ tên'] = df.iloc[:, 4]        # Cột E
    df_kq['Mã bệnh nhân'] = df.iloc[:, 3]  # Cột D
    df_kq['Năm sinh'] = df.apply(get_namsinh, axis=1)
    df_kq['Giới tính'] = df.apply(get_gioitinh, axis=1)

    # Lọc mã đúng định dạng BN******
    df_kq = df_kq[df_kq['Mã bệnh nhân'].astype(str).str.match(r'^BN\d{8}$')]

    # Gán chuyên khoa đúng cột và giữ xuống dòng
    df_kq['Thể lực'] = df.iloc[:, 8]       # Cột I
    df_kq['Nội khoa'] = df.iloc[:, 20]     # Cột U
    df_kq['Ngoại'] = df.iloc[:, 17]        # Cột R
    df_kq['Da liễu'] = df.iloc[:, 15]      # Cột P
    df_kq['Mắt'] = df.iloc[:, 10]          # Cột K
    df_kq['Tai mũi họng'] = df.iloc[:, 11] # Cột L
    df_kq['Răng hàm mặt'] = df.iloc[:, 13] # Cột N
    df_kq['Sản phụ khoa'] = df.iloc[:, 18]     # Cột S

    # ==== Tách và chèn cột thể lực ====
    def extract_data_from_theluc(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return pd.Series([None, None, None, None, None], 
                            index=['Chiều cao (cm)', 'Cân nặng (kg)', 'BMI', 'Mạch (l/ph)', 'Huyết áp(mmHg)'])

        chieu_cao = re.search(r'Chiều cao\s*:\s*([\d.]+)\s*cm', text)
        can_nang = re.search(r'Cân nặng\s*:\s*([\d.]+)\s*kg', text)
        bmi = re.search(r'BMI\s*:\s*([\d.]+)', text)
        mach = re.search(r'Mạch\s*:\s*([\d.]+)\s*l/p', text)
        huyet_ap = re.search(r'Huyết áp\s*:\s*([\d/]+)', text)

        return pd.Series([
            chieu_cao.group(1) if chieu_cao else None,
            can_nang.group(1) if can_nang else None,
            bmi.group(1) if bmi else None,
            mach.group(1) if mach else None,
            huyet_ap.group(1) if huyet_ap else None
        ], index=['Chiều cao', 'Cân nặng', 'BMI', 'Mạch', 'Huyết áp'])

    # Tách dữ liệu thể lực
    df_theluc = df_kq['Thể lực'].apply(extract_data_from_theluc)

    # Xóa cột "Thể lực"
    df_kq.drop(columns=['Thể lực'], inplace=True)

    # Chèn cột sau "Giới tính"
    cols = list(df_kq.columns)
    insert_at = cols.index('Giới tính') + 1
    for i, col_name in enumerate(['Chiều cao', 'Cân nặng', 'BMI', 'Mạch', 'Huyết áp']):
        df_kq.insert(loc=insert_at + i, column=col_name, value=df_theluc[col_name])


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Da liễu" và "Ngoại" ===========
    def classify_binh_thuong_or_bat_thuong(text):
        if text is None or pd.isna(text):
            return ""
    
        cleaned_text = str(text).strip().lower()
        
        # Nếu sau khi làm sạch mà rỗng hoặc là "nan", "none", "không khám", thì coi như không có thông tin
        if cleaned_text in ["", "nan", "none", "không khám", "chưa khám"]:
            return ""

        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
            "chưa ghi nhận bệnh lý"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        if any(keyword in cleaned_text for keyword in normal_keywords):
            return "BT"  # "BT" cho Bình thường
        else:
            return "TD"
        

    # Áp dụng hàm phân loại cho cột "Da liễu" và "Ngoại" và thay thế dữ liệu trong các cột đó
    df_kq['Da liễu'] = df_kq['Da liễu'].apply(classify_binh_thuong_or_bat_thuong)
    df_kq['Ngoại'] = df_kq['Ngoại'].apply(classify_binh_thuong_or_bat_thuong)

    def tong_hop_ngoai_khoa_da_lieu(row):
        da_lieu = row['Da liễu']
        ngoai = row['Ngoại']

        if da_lieu == 'BT' and ngoai == 'BT':
            return 'BT'
        elif da_lieu == 'TD' or ngoai == 'TD':
            return 'TD'
        else:
            return ''  # Không có dữ liệu

    df_kq['Ngoại khoa, da liễu'] = df_kq.apply(tong_hop_ngoai_khoa_da_lieu, axis=1)


    # Xóa 2 cột "Da liễu" và "Ngoại" cũ
    df_kq.drop(columns=['Da liễu', 'Ngoại'], inplace=True)

    # Di chuyển cột "Ngoại khoa, da liễu" ngay sau "Nội khoa"
    cols = df_kq.columns.tolist()
    cols.remove('Ngoại khoa, da liễu')
    noi_khoa_index = cols.index('Nội khoa') + 1
    cols.insert(noi_khoa_index, 'Ngoại khoa, da liễu')
    df_kq = df_kq[cols]


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột 'Tai mũi họng' ===========
    def classify_tai_mui_hong(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""
        
        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT" 
        
        # Danh sách từ khóa "Bệnh lý" (chỉ ví dụ, có thể mở rộng thêm)
        abnormal_keywords = [
            "viêm", 
            "nhiễm", 
            "bệnh lý", 
            "đau", 
            "chảy mủ", 
            "sưng"
        ]
        
        # Kiểm tra nếu có chứa ít nhất một từ khóa bất thường
        if any(keyword in text for keyword in abnormal_keywords):
            return "TD"  # "TD" cho Bất thường
        else:   
            return "BT"  # Nếu không có gì bất thường, mặc định là "BT"

    # Áp dụng hàm phân loại cho cột "Tai mũi họng" và thay thế dữ liệu trong cột đó
    df_kq['Tai mũi họng'] = df_kq['Tai mũi họng'].apply(classify_tai_mui_hong)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột 'Răng hàm mặt' ===========
    def classify_rang_ham_mat(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""
        
        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "viêm nướu mãn", 
            "mòn răng"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT"  # "BT" cho Bình thường
        
        # Danh sách từ khóa "Bệnh lý" (chỉ ví dụ, có thể mở rộng thêm)
        abnormal_keywords = [
            "viêm", 
            "nhiễm", 
            "bệnh lý", 
            "đau", 
            "sưng", 
            "lở", 
            "chảy máu", 
            "sâu"
        ]
        
        # Kiểm tra nếu có chứa ít nhất một từ khóa bất thường
        if any(keyword in text for keyword in abnormal_keywords):
            return "TD" 
        else:
            return "BT" 

    # Áp dụng hàm phân loại cho cột "Răng hàm mặt" và thay thế dữ liệu trong cột đó
    df_kq['Răng hàm mặt'] = df_kq['Răng hàm mặt'].apply(classify_rang_ham_mat)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Sản phụ khoa" ===========
    def classify_phu_khoa(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Danh sách từ khóa "Bình thường" cho Sản phụ khoa
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
            "chưa ghi nhận bệnh lý"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT"
        else:
            return "TD"

    # Áp dụng hàm phân loại cho cột "Sản phụ khoa" và thay thế trực tiếp dữ liệu trong cột đó
    df_kq['Sản phụ khoa'] = df_kq['Sản phụ khoa'].apply(classify_phu_khoa)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Nội khoa" ===========
    def classify_noikhoa(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Các tiểu mục bình thường của nội khoa
        normal_keywords = [
            "chưa thấy bất thường", "chưa ghi nhận bất thường", "chưa ghi nhận bệnh lý"
        ]

        # Danh sách tiểu mục kiểm tra
        tiemuc_keywords = [
            "tim mạch", "hô hấp", "tiêu hóa", "thận", "thần kinh", "nội tiết"
        ]

        # Kiểm tra nếu tất cả các tiểu mục đều bình thường (có chứa ít nhất một trong các từ khóa "bình thường")
        all_normal = all(any(normal in text.lower() for normal in normal_keywords) for keyword in tiemuc_keywords)

        if all_normal:
            return "BT"  # "BT" cho Bình thường
        else:
            return "TD"

    # Áp dụng hàm phân loại cho cột "Nội khoa" và thay thế trực tiếp dữ liệu trong cột đó
    df_kq['Nội khoa'] = df_kq['Nội khoa'].apply(classify_noikhoa)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Mắt" ===========
    def classify_mat(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Danh sách từ khóa xác định là bình thường
        normal_keywords = [
            "chưa thấy bất thường",
            "chưa ghi nhận bệnh lý",
            "không ghi nhận bất thường"
        ]

        text_lower = text.lower()
        for keyword in normal_keywords:
            if keyword in text_lower:
                return "BT"
        
        return "TD"

    # Áp dụng vào cột 'Mắt'
    df_kq['Mắt'] = df_kq['Mắt'].apply(classify_mat)


    # Lấy dữ liệu từ cột 30 (AE), loại bỏ đầu dòng là "-" hoặc "--" và strip() hai đầu
    def clean_issues(text):
        if pd.isna(text):
            return ""
        lines = str(text).splitlines()
        cleaned = []
        for line in lines:
            line = line.strip()
            # Nếu dòng bắt đầu bằng bất kỳ ký hiệu liệt kê nào, thay thế thành '-'
            # Bao gồm: --, -, +, *, •, v.v.
            line = re.sub(r'^[\-\+\*•]{1,}', '-', line)
            # Nếu sau dấu - có khoảng trắng thì giữ lại, nếu không thì thêm khoảng trắng
            line = re.sub(r'^-(\S)', r'- \1', line)
            cleaned.append(line)
        # Loại các dòng rỗng nếu có
        cleaned = [l for l in cleaned if l]
        return '\n'.join(cleaned)
    
    # Áp dụng hàm làm sạch cho cột 30 (AE)
    df_kq['Các vấn đề cần lưu ý'] = df.iloc[:, 30].apply(clean_issues)


    # ======= Phân loại sức khỏe =======
    # Cột Y (24), Z (25), AA (26), AB (27), AC (28)
    phanloai_map = {
        24: 'Loại I',
        25: 'Loại II',
        26: 'Loại III',
        27: 'Loại IV',
        28: 'Loại V',
    }

    def get_phanloai(row):
        for col_idx, label in phanloai_map.items():
            val = str(row[col_idx]).strip().lower()
            if val == 'x':
                return label
        return None

    df_kq['Phân loại sức khỏe'] = df.apply(get_phanloai, axis=1)
    
    # Lưu lại vào pickle
    # save_df(df_kq, "clinical_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df, out_pickle)


def run_py3(imaging_list_file, out_pickle):
    # Lấy năm hiện tại
    current_year = datetime.now().year

    # Đọc file, bỏ qua 16 dòng đầu
    df_raw = pd.read_excel(imaging_list_file, header=None, skiprows=16)
    # Chuẩn bị danh sách dữ liệu hợp lệ
    data = []

    for idx, row in df_raw.iterrows():
        # Kiểm tra Mã BA hợp lệ
        ma_ba = str(row[3]) if pd.notna(row[3]) else ""
        if not re.fullmatch(r"BA\d{8}", ma_ba):
            continue

        # Lấy họ tên từ cột G-H-I
        ho_ten = row[6] if pd.notna(row[6]) else row[7] if pd.notna(row[7]) else row[8]

        # Tuổi + Giới tính -> chuyển sang Năm sinh
        tuoi_nam = row[9]
        tuoi_nu = row[10]
        gioi_tinh = ""
        nam_sinh = ""

        if pd.notna(tuoi_nam):
            gioi_tinh = "Nam"
            nam_sinh = current_year - int(tuoi_nam)
        elif pd.notna(tuoi_nu):
            gioi_tinh = "Nữ"
            nam_sinh = current_year - int(tuoi_nu)

        # Yêu cầu và kết quả
        yeu_cau = row[18] if pd.notna(row[18]) else row[19]
        ket_qua = row[20]

        # Nếu có đủ họ tên và yêu cầu thì thêm vào danh sách
        if pd.notna(ho_ten) and pd.notna(yeu_cau):
            data.append({
                "Mã BA": ma_ba,
                "Họ tên người bệnh": ho_ten,
                "Giới tính": gioi_tinh,
                "Năm sinh": nam_sinh,
                "Yêu cầu": yeu_cau,
                "Kết quả": ket_qua
            })

    # Tạo DataFrame từ danh sách đã xử lý
    df = pd.DataFrame(data)

    # Lưu lại vào pickle
    # save_df(df, "imaging_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df, out_pickle)


def run_py4(clinical_temp_file, test_temp_file, out_pickle):
    """
    clinical_temp_file, test_temp_file: file-like object, đã mở ở chế độ 'rb'
    """
    unexpected_data_list = []
    # Load từ pickle
    df1 = pickle.load(clinical_temp_file)
    df2 = pickle.load(test_temp_file)

    def normalize_string(s):
        return str(s).strip().upper()

    df1['Họ tên chuẩn hóa'] = df1['Họ tên'].apply(normalize_string)
    df1['Giới tính chuẩn hóa'] = df1['Giới tính'].apply(normalize_string)
    df1['Năm sinh'] = df1['Năm sinh'].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")

    df2['Họ tên chuẩn hóa'] = df2['Họ tên người bệnh'].apply(normalize_string)
    df2['Giới tính chuẩn hóa'] = df2['Giới tính'].apply(normalize_string)
    df2['Năm sinh'] = df2['Năm sinh'].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")

    ds_yeu_cau = [
        'WBC', 'NEU %', 'RBC', 'HGB', 'MCV', 'MCH', 'PLT',
        'Định lượng Ure', 'Định lượng Creatinin', 'eGFR (CKD-EPI)',
        'Protein', 'Leukocytes', 'Nitrites', 'Blood',
        'Định lượng Glucose',
        'Đo hoạt độ AST (GOT)', 'Đo hoạt độ ALT (GPT)', 'Đo hoạt độ GGT',
        'Định lượng Triglycerid', 'Định lượng HDL-C', 'Định lượng LDL-C',
        'Định lượng Acid Uric', 'Bạch cầu', 'Nấm', 'Trichomonas Vaginalis'
    ]

    # df2_filtered = df2[df2['Yêu cầu'].isin(ds_yeu_cau)]

    def safe_value(x):
        allowed_strings = {'Negative', 'Positive', '1+', '2+', '3+', '4+', '+', '++', '+++', '++++'}
        if pd.isna(x):
            return ""

        # Nếu là chuỗi dạng số có dấu nháy (ví dụ: "'123.4")
        if isinstance(x, str):
            x = x.strip().lstrip("'")  # bỏ leading space và dấu nháy '
            
            # Nếu sau xử lý mà chuỗi là số thì convert
            if re.fullmatch(r'^-?\d+(\.\d+)?$', x):
                return float(x)

            # Nếu thuộc allowed_strings
            if x.upper() in allowed_strings:
                return x.upper()

            return ''

        # Nếu là số thực sự
        try:
            float(x)
            return x
        except:
            return ''
    df2_filtered = df2[df2['Yêu cầu'].isin(ds_yeu_cau)].copy()

    pivot_df = df2_filtered.pivot_table(
        index=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa', 'Năm sinh'],
        columns='Yêu cầu',
        values='Kết quả',
        aggfunc='first'
    ).reset_index()

    merged_df = pd.merge(
        df1,
        pivot_df,
        how='left',
        on=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa', 'Năm sinh']
    )

    def safe_float(x):
        try:
            if pd.isna(x) or x is None:
                return None
            if isinstance(x, str):
                x = x.strip().lstrip("'")  # bỏ dấu nháy và khoảng trắng
                if x.upper() in ['Negative', 'Positive', '1+', '2+', '3+', '4+', '5+', '+', '++', '+++', '++++', '+++++', '+-']:
                    return None
            return float(x) if x else None
        except:
            return None


    def soi_tuoi_huyet_trang(row):
        gender = str(row.get('Giới tính', '')).strip().upper()
        pk = str(row.get('Sản phụ khoa', '')).strip().upper()
        name = row.get("Họ tên", "Unknown")

        local_unexpected = []

        # Nếu Sản phụ khoa đã kết luận là TD thì giữ nguyên
        if pk == "TD":
            return "TD"

        # Lấy dữ liệu chỉ số
        bc_raw = row.get('Bạch cầu', None)
        nam_raw = row.get('Nấm', None)
        tv_raw = row.get('Trichomonas Vaginalis', None)

        # === Nếu là NAM ===
        if 'NAM' in gender:
            if all(pd.isna(v) or str(v).strip() == '' for v in [bc_raw, nam_raw, tv_raw]):
                return ""
            else:
                local_unexpected.append((name, "Soi tươi (nam nhưng có dữ liệu)", f"Bạch cầu: {bc_raw}, Nấm: {nam_raw}, TV: {tv_raw}"))
                unexpected_data_list.extend(local_unexpected)
                return ""

        # === Nếu là NỮ ===
        try:
            # Nếu tất cả đều trống → return ""
            if all(pd.isna(v) or str(v).strip() == '' for v in [bc_raw, nam_raw, tv_raw]):
                return ""

            # Bạch cầu
            bc_abnormal = False
            if bc_raw and str(bc_raw).strip() != "":
                try:
                    bc_str = str(bc_raw).strip().upper()
                    if '-' in bc_str:
                        parts = bc_str.split('-')
                        bc_val = float(parts[-1])
                    else:
                        bc_val = float(bc_str)
                    if bc_val > 20:
                        bc_abnormal = True
                except:
                    local_unexpected.append((name, "Bạch cầu", bc_raw))

            # Nấm
            nam_abnormal_values = ['1+', '2+', '3+', '4+', '+', '++', '+++', '++++']
            nam_normal_values = ['NEGATIVE']
            nam_str = str(nam_raw).strip().upper()
            nam_abnormal = nam_str in nam_abnormal_values
            if nam_str and nam_str not in nam_abnormal_values + nam_normal_values:
                local_unexpected.append((name, "Nấm", nam_str))

            # Trichomonas Vaginalis
            tv_str = str(tv_raw).strip().upper()
            tv_abnormal = tv_str == 'POSITIVE'
            if tv_str and tv_str not in ['POSITIVE', 'NEGATIVE']:
                local_unexpected.append((name, "Trichomonas Vaginalis", tv_str))

            # Ghi lại các dữ liệu không chuẩn
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Kết luận
            if bc_abnormal or nam_abnormal or tv_abnormal:
                return "TD"
            else:
                return "BT"

        except Exception as e:
            unexpected_data_list.append((name, "Soi tươi", f"Exception: {str(e)}"))
            return ""


    def tong_phan_tich(row):
        keys = ['WBC', 'NEU %', 'RBC', 'MCV', 'PLT']
        local_unexpected = []
        name = row.get("Họ tên", "Unknown")

        try:
            # Tách từng chỉ số và kiểm tra dữ liệu
            values = {}
            has_any_data = False  # Cờ để kiểm tra có dữ liệu hợp lệ nào không

            for k in keys:
                val = row.get(k, None)
                if val is None or pd.isna(val) or str(val).strip() == "":
                    values[k] = None
                else:
                    try:
                        values[k] = float(val)
                        has_any_data = True
                    except ValueError:
                        local_unexpected.append((name, k, f"Không phải số: {val}"))
                        values[k] = None

            # Nếu không có dữ liệu hợp lệ nào → return ""
            if not has_any_data:
                return ""

            # Nếu có lỗi kiểu dữ liệu → lưu lại nhưng vẫn tiếp tục đánh giá những chỉ số còn lại
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Phân tích giá trị
            def classify_huyet_hoc(values):
                # WBC
                wbc = values.get('WBC')
                if wbc is not None and not (2 < wbc < 14):
                    return "TD"

                # NEU %
                neu = values.get('NEU %')
                if neu is not None and neu >= 90:
                    return "TD"

                # RBC
                rbc = values.get('RBC')
                if rbc is not None and not (3.5 < rbc < 6):
                    return "TD"

                # MCV
                mcv = values.get('MCV')
                if mcv is not None and mcv <= 80:
                    return "TD"

                # PLT
                plt = values.get('PLT')
                if plt is not None and not (100 < plt < 450):
                    return "TD"

                # Nếu có ít nhất một giá trị hợp lệ và tất cả đều trong ngưỡng
                return "BT"

            return classify_huyet_hoc(values)

        except Exception as e:
            unexpected_data_list.append((name, "Tổng phân tích", f"Exception: {str(e)}"))
            return ""


    def men_gan(row):
        keys = ['Đo hoạt độ AST (GOT)', 'Đo hoạt độ ALT (GPT)']
        local_unexpected = []
        name = row.get("Họ tên", "Unknown")

        try:
            ast_val = row.get(keys[0], None)
            alt_val = row.get(keys[1], None)

            # Kiểm tra AST
            if pd.isna(ast_val) or str(ast_val).strip() == "":
                local_unexpected.append((name, keys[0], "Dữ liệu trống"))
                ast = None
            else:
                try:
                    ast = float(ast_val)
                except ValueError:
                    local_unexpected.append((name, keys[0], f"Không phải số: {ast_val}"))
                    ast = None

            # Kiểm tra ALT
            if pd.isna(alt_val) or str(alt_val).strip() == "":
                local_unexpected.append((name, keys[1], "Dữ liệu trống"))
                alt = None
            else:
                try:
                    alt = float(alt_val)
                except ValueError:
                    local_unexpected.append((name, keys[1], f"Không phải số: {alt_val}"))
                    alt = None

            # Nếu có dữ liệu bất thường, lưu lại
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Xử lý kết luận
            if ast is not None and ast >= 90:
                return "TD"
            if alt is not None and alt >= 90:
                return "TD"

            # Nếu có ít nhất một giá trị hợp lệ và không bất thường
            if ast is not None or alt is not None:
                return "BT"

            # Nếu cả hai giá trị đều thiếu
            return ""

        except Exception as e:
            unexpected_data_list.append((name, "Men gan", f"Exception: {str(e)}"))
            return ""


    def chuc_nang_than(row):
        key = 'eGFR (CKD-EPI)'
        local_unexpected = []

        try:
            val = row.get(key, None)
            name = row.get("Họ tên", "Unknown")

            if pd.isna(val) or str(val).strip() == "":
                local_unexpected.append((name, key, "Dữ liệu trống"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            try:
                val_float = float(val)
            except ValueError:
                local_unexpected.append((name, key, f"Không phải số: {val}"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            return "BT" if val_float > 80 else "TD"

        except Exception as e:
            unexpected_data_list.append((name, key, f"Exception: {str(e)}"))
            return ""


    def duong_huyet(row):
        key = 'Định lượng Glucose'
        local_unexpected = []

        try:
            val = row.get(key, None)
            name = row.get("Họ tên", "Unknown")

            if pd.isna(val) or str(val).strip() == "":
                local_unexpected.append((name, key, "Dữ liệu trống"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            try:
                val_float = float(val)
            except ValueError:
                local_unexpected.append((name, key, f"Không phải số: {val}"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            # Nếu hợp lệ, kiểm tra giá trị
            return "BT" if 60 <= val_float <= 129 else "TD"
        
        except Exception as e:
            unexpected_data_list.append((name, key, f"Exception: {str(e)}"))
            return ""


    def lipid(row):
        keys = ['Định lượng Triglycerid', 'Định lượng LDL-C', 'Định lượng HDL-C', 'Định lượng Cholesterol']
        local_unexpected = []
        results = {}  # Lưu trạng thái từng chỉ số: 'TD', 'BT', 'TRONG'
        name = row.get("Họ tên", "Unknown")

        try:
            for key in keys:
                val = row.get(key, None)
                if pd.isna(val) or str(val).strip() == "":
                    results[key] = 'TRONG'
                    local_unexpected.append((name, key, "Dữ liệu trống"))
                    continue
                try:
                    val_float = safe_float(val)
                except:
                    results[key] = 'TRONG'
                    local_unexpected.append((name, key, f"Không phải số: {val}"))
                    continue

                # Xét kết quả cho từng chỉ số
                if key == 'Định lượng Triglycerid':
                    results[key] = 'TD' if val_float >= 150 else 'BT'
                elif key == 'Định lượng LDL-C':
                    results[key] = 'TD' if val_float >= 129 else 'BT'
                elif key == 'Định lượng HDL-C':
                    results[key] = 'TD' if val_float <= 40 else 'BT'
                elif key == 'Định lượng Cholesterol':
                    results[key] = 'TD' if val_float > 200 else 'BT'

            # Nếu có bất thường dữ liệu
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            status_list = list(results.values())
            if 'TD' in status_list:
                return 'TD'
            elif all(st == 'BT' for st in status_list):
                return 'BT'
            elif 'BT' in status_list and 'TD' not in status_list:
                return 'BT'
            else:
                return ''  # Trường hợp toàn bộ là trống hoặc bất thường

        except Exception as e:
            unexpected_data_list.append((name, "Exception", str(e)))
            return ""


    def sinh_hoa_khac(row):
        key = 'Định lượng Acid Uric'
        try:
            val = safe_float(row[key])
            name = row.get("Họ tên", "Unknown")
            if val is None:
                return ""
            gender = str(row['Giới tính']).strip().upper()
            if ('NAM' in gender and val >= 420) or ('NỮ' in gender and val >= 360):
                return "TD"
            elif 'NAM' in gender and val < 420:
                return 'BT'
            elif 'NỮ' in gender and val < 360:
                return 'BT'
            
            return ""
        except Exception as e:
            unexpected_data_list.append((name, "Exception", str(e)))
            return ""


    def nuoc_tieu(row):
        keys = ['Protein', 'Leukocytes', 'Nitrites', 'Blood']
        abnormal_values = ['POSITIVE', '2+', '3+', '4+', '5+', '6+', '7+', '++', '+++', '++++', '+++++', '++++++', '+++++++']
        normal_values = ['NEGATIVE', '1+', '0', '+-']
        unexpected_data = []
        values = []
        abnormal_found = False
        empty_count = 0  # Đếm số ô trống

        try:
            for k in keys:
                val = row.get(k, None)
                val_str = str(val).strip().upper() if val is not None else ""

                if val_str == "" or val_str == "NAN":
                    empty_count += 1
                    continue

                if k in ['Leukocytes', 'Blood']:
                    try:
                        num_val = float(val_str)
                        if k == 'Blood' and num_val >= 80:
                            return "TD"
                        elif k == 'Leukocytes' and num_val > 100:
                            return "TD"
                        else:
                            values.append("BT")
                            continue
                    except Exception:
                        pass  # không phải số, xử lý tiếp như chuỗi

                if val_str in abnormal_values:
                    abnormal_found = True
                elif val_str in normal_values:
                    values.append("BT")
                else:
                    unexpected_data.append((row.get('Họ tên', 'Không rõ'), k, val))
            
            if unexpected_data:
                unexpected_data_list.extend(unexpected_data)

            if empty_count == len(keys):
                return ""

            if abnormal_found:
                return "TD"

            if values and all(v == "BT" for v in values):
                return "BT"

            return ""  # nếu không rõ ràng thì để trống

        except Exception as e:
            unexpected_data_list.append((row.get('Họ tên', 'Không rõ'), 'nuoc_tieu', str(e)))
            return ""


        
    merged_df['Soi tươi huyết trắng'] = merged_df.apply(soi_tuoi_huyet_trang, axis=1)
    merged_df['Tổng phân tích tế bào máu ngoại vi'] = merged_df.apply(tong_phan_tich, axis=1)
    merged_df['Men gan'] = merged_df.apply(men_gan, axis=1)
    merged_df['Chức năng thận'] = merged_df.apply(chuc_nang_than, axis=1)
    merged_df['Đường huyết lúc đói'] = merged_df.apply(duong_huyet, axis=1)
    merged_df['Xét nghiệm lipid máu'] = merged_df.apply(lipid, axis=1)
    merged_df['Sinh hóa khác'] = merged_df.apply(sinh_hoa_khac, axis=1)
    merged_df['Tổng phân tích nước tiểu'] = merged_df.apply(nuoc_tieu, axis=1)

    merged_df.drop(columns=[col for col in ds_yeu_cau if col in merged_df.columns], inplace=True)
    merged_df.drop(columns=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa'], inplace=True)

    # ## Đảm bảo 'Các vấn đề cần lưu ý' và 'Phân loại sức khỏe' nằm cuối cùng
    # cols = [col for col in merged_df.columns if col not in ['Các vấn đề cần lưu ý', 'Phân loại sức khỏe']]
    # if 'Các vấn đề cần lưu ý' in merged_df.columns:
    #     cols.append('Các vấn đề cần lưu ý')
    # if 'Phân loại sức khỏe' in merged_df.columns:
    #     cols.append('Phân loại sức khỏe')
    # merged_df = merged_df[cols]

    # Kiểm tra lỗi chính tả và thêm vào unexpected_data_list nếu có
    # if 'Các vấn đề cần lưu ý' in merged_df.columns:
    #     for idx, row in merged_df.iterrows():
    #         text = row['Các vấn đề cần lưu ý']
    #         loi_chinh_ta = kiem_tra_chinh_ta_viet(text)
    #         if loi_chinh_ta:
    #             ho_ten = row.get('Họ tên', row.get('Họ tên chuẩn hóa', 'Không rõ'))
    #             loi_text = ", ".join(loi_chinh_ta)
    #             unexpected_data_list.append((ho_ten, 'Chính tả', f'Lỗi: {loi_text}'))

    # Lưu lại vào pickle
    # save_df(merged_df, "merge_clinical_test_temp.pkl")
    # return unexpected_data_list

    # Lưu merged_df ra file-like object pickle
    pickle.dump(merged_df, out_pickle)
    return unexpected_data_list


# def run_py5(merged_file, clinical_list_file, imaging_temp_file, customer_list_file, output_file):
#     """
#     merged_file: file-like object (pickle, 'rb')
#     clinical_list_file: file-like object (excel, 'rb')
#     imaging_temp_file: file-like object (pickle, 'rb')
#     customer_list_file: file-like object (excel, 'rb')
#     output_file: path hoặc file-like object ('wb')
#     """
#     # Load từ pickle
#     df_base = pickle.load(merged_file)
#     df_cdha = pickle.load(imaging_temp_file)

#     # ===== Chuẩn hóa để ghép dữ liệu =====
#     df_base['Họ tên'] = df_base['Họ tên'].astype(str).str.strip()
#     df_base['Giới tính'] = df_base['Giới tính'].astype(str).str.strip()
#     df_base['Năm sinh'] = df_base['Năm sinh'].astype(str).str.strip()

#     df_cdha['Họ tên'] = df_cdha['Họ tên người bệnh'].astype(str).str.strip()
#     df_cdha['Giới tính'] = df_cdha['Giới tính'].astype(str).str.strip()
#     df_cdha['Năm sinh'] = df_cdha['Năm sinh'].astype(str).str.strip()

#     # ===== Lọc các dịch vụ cần lấy =====
#     dich_vu_chon = [
#         'Chụp X-quang Ngực thẳng',
#         'Siêu âm tuyến giáp',
#         'Siêu âm tuyến vú hai bên'
#     ]

#     # Lọc các hàng có "Siêu âm ổ bụng" trong cột 'Yêu cầu' và các dịch vụ khác
#     df_cdha_filtered = df_cdha[
#         df_cdha['Yêu cầu'].isin(dich_vu_chon) |
#         df_cdha['Yêu cầu'].str.contains('Siêu âm ổ bụng', na=False)
#     ]

#     # Chuẩn hóa tên cột 'Siêu âm ổ bụng' trong df_cdha_filtered
#     df_cdha_filtered.loc[df_cdha_filtered['Yêu cầu'].str.contains('Siêu âm ổ bụng', na=False), 'Yêu cầu'] = 'Siêu âm ổ bụng'

#     # ===== Pivot để mỗi dịch vụ là một cột =====
#     df_pivot = df_cdha_filtered.pivot_table(
#         index=['Họ tên', 'Giới tính', 'Năm sinh'],
#         columns='Yêu cầu',
#         values='Kết quả',
#         aggfunc='first'
#     ).reset_index()

#     # ===== Ghép dữ liệu CDHA vào dữ liệu gốc =====
#     df_merged = pd.merge(df_base, df_pivot, on=['Họ tên', 'Giới tính', 'Năm sinh'], how='left')

#     # ===== Hàm đánh giá BT/TD cho các cột =====
#     def danh_gia(text, dieu_kien_bt):
#         if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
#             return ''
#         val_lower = str(text).lower()
#         for dk in dieu_kien_bt:
#             if dk in val_lower:
#                 return 'BT'
#         return 'TD'

#     # ===== Hàm đánh giá riêng cho Siêu âm ổ bụng =====
#     def danh_gia_sieu_am_o_bung(val):
#         if pd.isna(val):
#             return ''
#         val_lower = str(val).lower()
#         if 'chưa phát hiện bất thường' in val_lower:
#             return 'BT'
#         return 'TD'

#     # ===== Đánh giá và thay thế trực tiếp các kết quả CDHA =====
#     cot_danh_gia = {
#         'Chụp X-quang Ngực thẳng': ['chưa thấy bất thường', 'trong giới hạn bình thường'],
#         'Siêu âm tuyến giáp': ['chưa thấy bất thường', 'chưa phát hiện bất thường'],
#         'Siêu âm tuyến vú hai bên': ['chưa thấy bất thường', 'chưa phát hiện bất thường']
#     }

#     for col, dieu_kien_bt in cot_danh_gia.items():
#         if col in df_merged.columns:
#             df_merged[col] = df_merged[col].apply(lambda x: danh_gia(x, dieu_kien_bt))

#     # ===== Đánh giá và thay thế trực tiếp cho Siêu âm ổ bụng =====
#     if 'Siêu âm ổ bụng' in df_merged.columns:
#         df_merged['Siêu âm ổ bụng'] = df_merged['Siêu âm ổ bụng'].apply(danh_gia_sieu_am_o_bung)

#     # ===== Đảm bảo hai cột 'Các vấn đề cần lưu ý' và 'Phân loại sức khỏe' nằm cuối =====
#     cols = [col for col in df_merged.columns if col not in ['Các vấn đề cần lưu ý', 'Phân loại sức khỏe']]
#     if 'Các vấn đề cần lưu ý' in df_merged.columns:
#         cols.append('Các vấn đề cần lưu ý')
#     if 'Phân loại sức khỏe' in df_merged.columns:
#         cols.append('Phân loại sức khỏe')
#     df_merged = df_merged[cols]

#     # =============================== Cập nhật ngày sinh từ DS_KSK.xlsx và tô màu lỗi ======================================#
#     # Đọc file DS_KSK.xlsx
#     df_ksk = pd.read_excel(customer_list_path, skiprows=1, dtype=str)

#     # Chuẩn hóa tên cột: xóa khoảng trắng, chữ thường
#     df_ksk.columns = df_ksk.columns.str.strip().str.lower()

#     # Lọc những dòng có mã BN đúng định dạng BN + 6 chữ số
#     pattern = re.compile(r'^BN\d{6}$')
#     df_filtered = df_ksk[df_ksk['mã bn'].apply(lambda x: bool(pattern.match(str(x).strip())))].copy()

#     # Chuẩn hóa lại dữ liệu
#     df_filtered['mã bn'] = df_filtered['mã bn'].str.strip()
#     df_filtered['họ tên'] = df_filtered['họ tên'].str.strip().str.lower()
#     df_filtered['giới tính'] = df_filtered['giới tính'].str.strip().str.capitalize()
#     df_filtered['ngày sinh'] = pd.to_datetime(df_filtered['ngày sinh'], errors='coerce')

#     # Thêm các cột xử lý bổ sung
#     df_filtered['ngày sinh dạng chuỗi'] = df_filtered['ngày sinh'].dt.strftime('%d/%m/%Y')
#     df_filtered['năm sinh'] = df_filtered['ngày sinh'].dt.year.astype('Int64')

#     # Tạo dict tra cứu theo (Họ Tên, Mã BN)
#     ksk_dict = {
#         (row['họ tên'], row['mã bn']): (row['ngày sinh dạng chuỗi'], row['năm sinh'], row['giới tính'])
#         for _, row in df_filtered.iterrows()
#     }

#     # Chuẩn hóa dữ liệu file 2
#     df_merged['Họ tên'] = df_merged['Họ tên'].astype(str).str.strip().str.lower()
#     df_merged['Mã bệnh nhân'] = df_merged['Mã bệnh nhân'].astype(str).str.strip().str.lower()
#     df_merged['Năm sinh'] = df_merged['Năm sinh'].astype(str).str.strip()

#     # Cập nhật ngày sinh
#     def cập_nhật_ngày_sinh(row):
#         key = (row['Họ tên'], row['Mã bệnh nhân'])
#         if key in ksk_dict:
#             ngay_sinh, nam_sinh = ksk_dict[key]
#             try:
#                 if int(row['Năm sinh']) == int(nam_sinh):
#                     return ngay_sinh
#                 else:
#                     return "Lỗi năm sinh"
#             except:
#                 return "Lỗi năm sinh"
#         else:
#             return row['Năm sinh']  # giữ nguyên nếu không khớp

#     # Áp dụng cập nhật
#     df_merged['Năm sinh'] = df_merged.apply(cập_nhật_ngày_sinh, axis=1)
#     # ⭐ Chuyển Họ tên và Mã bệnh nhân về in hoa sau cập nhật
#     df_merged['Họ tên'] = df_merged['Họ tên'].str.upper()
#     df_merged['Mã bệnh nhân'] = df_merged['Mã bệnh nhân'].str.upper()

#     #================================================================================#

#     # ===== THÊM CỘT STT =====
#     df_merged.insert(0, 'STT', range(1, len(df_merged) + 1))


#     # ===== Xuất ra file Excel =====
#     # output_file = Path(settings.MEDIA_ROOT) 
#     df_merged.to_excel(output_path, index=False)

#     # ===== Định dạng bảng Excel =====
#     wb = openpyxl.load_workbook(output_path)
#     ws = wb.active

#     # Style chung
#     thin_border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     header_font = Font(bold=True)

#     # Định dạng từng ô
#     for col in ws.columns:
#         max_length = 0
#         col_letter = get_column_letter(col[0].column)
#         for cell in col:
#             # Viền
#             cell.border = thin_border

#             if cell.row == 1:
#                 # Hàng tiêu đề
#                 cell.fill = header_fill
#                 cell.font = header_font
#                 cell.alignment = Alignment(
#                     text_rotation=90,
#                     vertical='center',
#                     horizontal='center',
#                     wrap_text=True
#                 )
#             else:
#                 value = str(cell.value).strip() if cell.value is not None else ''
#                 col_name = ws.cell(row=1, column=cell.column).value

#                 # Danh sách các cột cần căn giữa
#                 cac_cot_can_giua = [
#                     'STT', 'Mã bệnh nhân', 'Năm sinh','Giới tính', 'Chiều cao', 'Cân nặng', 'Huyết áp', 'BMI', 
#                     'Mạch', 'Nhiệt độ', 'Nhịp thở', 'Thị lực P', 'Thị lực T',
#                     'Phân loại sức khỏe'
#                 ]

#                 if value in ['BT', 'TD'] or col_name in cac_cot_can_giua:
#                     cell.alignment = Alignment(horizontal='center', vertical='center')
#                 elif col_name == 'Các vấn đề cần lưu ý':
#                     cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
#                 else:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')

#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         # Điều chỉnh độ rộng
#         if ws.cell(row=2, column=col[0].column).value in ['BT', 'TD']:
#             ws.column_dimensions[col_letter].width = 10
#         else:
#             ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))

#     # Chiều cao tiêu đề
#     ws.row_dimensions[1].height = 70

#     # xoay dọc mã BN và năm sinh
#     ## Bắt đầu từ dòng 2 (sau tiêu đề)
#     row = 2
#     while True:
#         cell_stt = ws.cell(row=row, column=1)  # Cột A

#         # Nếu cột A không phải là số → dừng
#         try:
#             if cell_stt.value is None or not isinstance(int(cell_stt.value), int):
#                 break
#         except:
#             break

#         # Áp dụng xoay dọc cho cột C (3) và D (4)
#         align = Alignment(textRotation=90, vertical='center', horizontal='center')
#         ws.cell(row=row, column=3).alignment = align  # Cột C: Mã BN
#         ws.cell(row=row, column=4).alignment = align  # Cột D: Năm sinh

#         row += 1
#     #=========================================== Tạo form Final=============================================#
#     # === Thêm 1 dòng đầu tiên ===
#     ws.insert_rows(1)

#     # Lấy tên công ty: giá trị ô dòng 6 (hàng số 6), cột C (cột 3)
#     df_raw_clinical = pd.read_excel(clinical_list_path, header=None) 
#     title_file = df_raw_clinical.iloc[5, 2]  # 5 = dòng 6, 2 = cột C

#     if title_file:
#         # Xóa phần đầu "DANH SÁCH KHÁM SỨC KHỎE ĐỊNH KỲ CBCNV " (41 ký tự)
#         company_name = title_file.replace("DANH SÁCH KHÁM SỨC KHỎE ĐỊNH KỲ CBCNV ", "").strip()
#         print("Tên công ty:", company_name)
#     else:
#         company_name = "UnknownCompany"
#         print("Không tìm thấy tên công ty")
#     safe_company_name = company_name.replace(" ", "_").replace("/", "_")  # tránh lỗi tên file
#     safe_company_name = slugify(company_name) 

#     # === Tạo tiêu đề kép trong một ô (dòng 1) ===
#     title_1 = "KẾT QUẢ KIỂM TRA SỨC KHỎE TỔNG QUÁT NĂM 2024"
#     title_2 = company_name 
#     full_title = f"{title_2}\n{title_1}"

#     max_col = ws.max_column

#     # Gộp toàn bộ dòng 1 từ A đến cột cuối
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

#     # Ghi nội dung + định dạng font và căn giữa
#     title_cell = ws.cell(row=1, column=1)
#     title_cell.value = full_title
#     title_cell.font = Font(name="Times New Roman", size=18, bold=True)
#     title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     # Thêm phần màu sắc
#     fill_formula_cell = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
#     fill_stat_header = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

#     # === Xác định số dòng dữ liệu (sau khi thêm 1 dòng tiêu đề) ===
#     start_data_row = 3
#     # === Tạo 7 dòng tổng hợp cuối bảng ===
#     summary_titles = [
#         "TỔNG SỐ THỰC HIỆN",
#         "TỔNG SỐ BÌNH THƯỜNG (BT)",
#         "TỔNG SỐ BẤT THƯỜNG CẦN THEO DÕI THÊM (TD)",
#         "TỔNG SỐ TỪ CHỐI (TC)",
#         "TỔNG SỐ KHÔNG THỰC HIỆN (Ô TRỐNG)",
#         "TỈ LỆ BẤT THƯỜNG CẦN THEO DÕI THÊM",
#         "TỔNG KHÁCH HÀNG"
#     ]

#     # Tính số dòng dữ liệu từ STT → dòng kết thúc thực sự
#     end_data_row = start_data_row
#     while True:
#         cell_stt = ws.cell(row=end_data_row, column=1)
#         try:
#             if cell_stt.value is None or not isinstance(int(cell_stt.value), int):
#                 break
#         except:
#             break
#         end_data_row += 1
#     end_data_row -= 1  # Trừ đi 1 vì dòng hiện tại là dòng không hợp lệ
#     summary_start_row = end_data_row + 1
#     summary_end_row = summary_start_row + len(summary_titles) - 1

#     # Duyệt từng dòng thống kê
#     for i, title in enumerate(summary_titles):
#         row = summary_start_row + i
#         # Gộp từ A đến J
#         ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
#         ws.cell(row=row, column=1, value=title).alignment = Alignment(horizontal="left", vertical="center")
#         cell = ws.cell(row=row, column=1, value=title)
#         cell.fill = fill_stat_header  # tô màu nền
        
#         for col in range(11, max_col):  # từ cột K (11) đến trước cột cuối cùng (AD)
#             col_letter = get_column_letter(col)
#             cell = ws.cell(row=row, column=col)
#             data_col_range = f"{col_letter}{start_data_row}:{col_letter}{end_data_row}"

#             if i == 0:  # Tổng số thực hiện
#                 cell.value = f"=COUNTA({data_col_range})"
#             elif i == 1:  # Đếm BT
#                 cell.value = f'=COUNTIF({data_col_range},"BT")'
#             elif i == 2:  # Đếm TD
#                 cell.value = f'=COUNTIF({data_col_range},"TD")'
#             elif i == 3:  # Đếm TC
#                 cell.value = f'=COUNTIF({data_col_range},"TC")'
#             elif i == 4:  # Đếm ô trống
#                 cell.value = f'=COUNTBLANK({data_col_range})'
#             elif i == 5:  # Tỷ lệ TD
#                 cell.value = f'=IFERROR(COUNTIF({data_col_range},"TD")/COUNTA({data_col_range}), "")'
#                 cell.number_format = '0 %'  # Hiển thị dưới dạng phần trăm, không chữ thập phân
#             elif i == 6:  # Tổng khách hàng (tính theo cột họ tên)
#                 cell.value = f'=COUNTA(B{start_data_row}:B{end_data_row})'
#             cell.alignment = Alignment(horizontal="center")
#             if col != 29 and col != 30:
#                 cell.fill = fill_formula_cell  # Đổ màu cho ô công thức

#     # === Merge và ghi ngày tháng + chữ ký bên phải (AD-AE) ===
#     last_col_letter = get_column_letter(max_col)
#     ws.merge_cells(start_row=summary_start_row, start_column=max_col - 1, end_row=summary_end_row, end_column=max_col)
#     ws.cell(row=summary_start_row, column=max_col - 1).value = "\n…, ngày … tháng … năm {today.year}\n\nGiám đốc Y khoa"
#     ws.cell(row=summary_start_row, column=max_col - 1).alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
#     ws.cell(row=summary_start_row, column=max_col - 1).font = Font(italic=True)

#     # === Kẻ viền cho 7 dòng cuối ===
#     thin = Side(style='thin')
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     for row in ws.iter_rows(min_row=summary_start_row, max_row=summary_end_row, min_col=1, max_col=max_col):
#         for cell in row:
#             cell.border = border
#     # ==================================
#     # filename = f"summary_report_{safe_company_name}.xlsx"
#     # output_path = os.path.join(settings.MEDIA_ROOT, filename)

#     wb.save(output_path)
#     wb.close()


from .process_py5 import run_py5


def run_full_pipeline(customer_list, test_list, clinical_list, imaging_list, output_path):
    """
    Các biến customer_list, test_list,... là file-like object (request.FILES['...']).
    output_path: path hoặc file-like object để xuất báo cáo.
    """
    # ==== Bước 1: Xử lý test_list, clinical_list, imaging_list ====
    # Tạo buffer memory để lưu pickle tạm
    test_temp = io.BytesIO()
    clinical_temp = io.BytesIO()
    imaging_temp = io.BytesIO()
    merge_clinical_test_temp = io.BytesIO()

    # run_py1, run_py2, run_py3 nhận file-like object (Excel),
    # lưu kết quả ra buffer pickle
    run_py1(test_list, out_pickle=test_temp)
    run_py2(clinical_list, out_pickle=clinical_temp)
    run_py3(imaging_list, out_pickle=imaging_temp)

    # Reset pointer để đọc lại pickle
    test_temp.seek(0)
    clinical_temp.seek(0)
    imaging_temp.seek(0)

    # ==== Bước 2: So sánh bất thường ====
    # run_py4 trả về pickle cho bước sau:
    merge_clinical_test_temp = io.BytesIO()
    unexpected_data_all = []
    unexpected_data_all.append(
        run_py4(clinical_temp, test_temp, out_pickle=merge_clinical_test_temp)
    )
    merge_clinical_test_temp.seek(0)

    # ==== Bước 3: Ghép dữ liệu & xuất báo cáo ====
    # run_py5 nhận pickle, excel file-like, trả file excel
    run_py5(
        merged_file=merge_clinical_test_temp,   # nếu cần
        clinical_list_file=clinical_list,
        imaging_temp_file=imaging_temp,
        customer_list_file=customer_list,
        output_file=output_path
    )

    return unexpected_data_all

# def run_full_pipeline(customer_list, test_list, clinical_list, imaging_list, output_path):

#     customer_list = Path("media/customer_list.xlsx")
#     imaging_list = Path("media/imaging_list.xlsx")
#     clinical_list = Path("media/clinical_list.xlsx")
#     test_list = Path("media/test_list.xlsx")

#     clinical_temp = Path("media/temp/clinical_temp.pkl")
#     test_temp = Path("media/temp/test_temp.pkl")
#     imaging_temp = Path("media/temp/imaging_temp.pkl")
#     merge_clinical_test_temp = Path("media/temp/merge_clinical_test_temp.pkl")

#     unexpected_data_all = []

#     run_py1(test_list)
#     run_py2(clinical_list)
#     run_py3(imaging_list)
#     unexpected_data_all.append(run_py4(clinical_temp, test_temp))
#     run_py5(merge_clinical_test_temp, clinical_list, imaging_temp, customer_list, output_path)

#     return unexpected_data_all

 
# Nếu muốn gọi thử riêng:
# if __name__ == "__main__":
#     output_path = Path(settings.MEDIA_ROOT) / "output_report"
#     run_full_pipeline('customer_list', 'test_list', 'clinical_list', 'imaging_list', 'output_path')
