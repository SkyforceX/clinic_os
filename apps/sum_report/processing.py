import os
from pathlib import Path
from django.conf import settings
from django.http import FileResponse
from django.utils.text import slugify
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import re
import pickle
from pyvi import ViTokenizer, ViUtils
import io

from .process_py5 import run_py5


def kiem_tra_chinh_ta_viet(text):
    # Tách từ
    tokens = ViTokenizer.tokenize(text).split()
    # Danh sách các từ không hợp lệ (ví dụ: chứa ký tự lạ, không trong từ điển unicode tiếng Việt)
    loi = []
    for token in tokens:
        word = token.replace("_", " ")
        # Kiểm tra xem có chứa ký tự không hợp lệ không
        if not re.match(r"^[a-zA-ZÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠàáâãèéêìíòóôõùúăđĩũơƯĂẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼỀỀỂưăạảấầẩẫậắằẳẵặẹẻẽềềểếỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪễệỉịọỏốồổỗộớờởỡợụủứừỬỮỰỲỴÝỶỸỳỵỷỹ\s\-]+$", word):
            loi.append(word)
    return loi


def run_py1(test_list_file, out_pickle):
    df_raw = pd.read_excel(test_list_file, header=None, skiprows=16)

    # Tìm dòng dừng khi cột E chứa 'NGƯỜI LẬP BIỂU'
    stop_index = df_raw[df_raw[4].astype(str).str.contains("NGƯỜI LẬP BIỂU", na=False)].index
    if not stop_index.empty:
        df_raw = df_raw.loc[:stop_index[0] - 1]

    # ======= Xử lý dữ liệu =======
    data = []
    for _, row in df_raw.iterrows():
        hoten = row[3]
        yeucau = row[15]
        ketqua = row[16]

        tuoi_nam = row[5]
        tuoi_nu = row[7]

        gioitinh = None
        tuoi = None

        if pd.notna(tuoi_nam) and isinstance(tuoi_nam, (int, float)):
            gioitinh = "Nam"
            tuoi = tuoi_nam
        elif pd.notna(tuoi_nu) and isinstance(tuoi_nu, (int, float)):
            gioitinh = "Nữ"
            tuoi = tuoi_nu

        namsinh = int(datetime.now().year - tuoi) if tuoi else None

        data.append({
            'Họ tên người bệnh': hoten,
            'Giới tính': gioitinh,
            'Năm sinh': namsinh,
            'Yêu cầu': yeucau,
            'Kết quả': ketqua
        })

    # ======= Tạo DataFrame và ghi ra file Excel =======
    df = pd.DataFrame(data)
    # Lọc dữ liệu hợp lệ
    df = df[df['Họ tên người bệnh'].apply(lambda x: isinstance(x, str) and x.strip() != '')]
    df = df[df['Giới tính'].isin(['Nam', 'Nữ'])]
    df = df[df['Năm sinh'].notna()]

    # Lưu ra file pickle
    # save_df(df, "test_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df, out_pickle)


def run_py2(clinical_list_file, out_pickle):
    # ======== Đọc và xử lý dữ liệu từ Excel ===========    
    df = pd.read_excel(clinical_list_file)

    # Hàm xác định giới tính
    def get_gioitinh(row):
        if pd.notna(row[5]):
            return 'Nam'
        elif pd.notna(row[7]):
            return 'Nữ'
        return None

    # Hàm lấy năm sinh tại đúng cột giới tính
    def get_namsinh(row):
        if pd.notna(row[5]):
            return row[5]
        elif pd.notna(row[7]):
            return row[7]
        return None

    # Tạo DataFrame kết quả
    df_kq = pd.DataFrame()
    df_kq['Họ tên'] = df.iloc[:, 4]        # Cột E
    df_kq['Mã bệnh nhân'] = df.iloc[:, 3]  # Cột D
    df_kq['Năm sinh'] = df.apply(get_namsinh, axis=1)
    df_kq['Giới tính'] = df.apply(get_gioitinh, axis=1)

    # Lọc mã đúng định dạng BN******
    df_kq = df_kq[df_kq['Mã bệnh nhân'].astype(str).str.match(r'^BN\d{8}$')]

    # Gán chuyên khoa đúng cột và giữ xuống dòng
    # =========== các cột cũ ===============
    # df_kq['Thể lực'] = df.iloc[:, 8]       # Cột I
    # df_kq['Nội khoa'] = df.iloc[:, 20]     # Cột U
    # df_kq['Ngoại'] = df.iloc[:, 17]        # Cột R
    # df_kq['Da liễu'] = df.iloc[:, 15]      # Cột P
    # df_kq['Mắt'] = df.iloc[:, 10]          # Cột K
    # df_kq['Tai mũi họng'] = df.iloc[:, 11] # Cột L
    # df_kq['Răng hàm mặt'] = df.iloc[:, 13] # Cột N
    # df_kq['Sản phụ khoa'] = df.iloc[:, 18]     # Cột S

    df_kq['Thể lực'] = df.iloc[:, 8]  # Cột I
    df_kq['Nội khoa'] = df.iloc[:, 21]  # Cột V
    df_kq['Ngoại'] = df.iloc[:, 18]  # Cột S
    df_kq['Da liễu'] = df.iloc[:, 16]  # Cột Q
    df_kq['Mắt'] = df.iloc[:, 10]  # Cột K
    df_kq['Tai mũi họng'] = df.iloc[:, 12]  # Cột M
    df_kq['Răng hàm mặt'] = df.iloc[:, 14]  # Cột O
    df_kq['Sản phụ khoa'] = df.iloc[:, 19]  # Cột T

    # ==== Tách và chèn cột thể lực ====
    def extract_data_from_theluc(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return pd.Series([None, None, None, None, None], 
                            index=['Chiều cao (cm)', 'Cân nặng (kg)', 'BMI', 'Mạch (l/ph)', 'Huyết áp(mmHg)'])

        chieu_cao = re.search(r'Chiều cao\s*:\s*([\d.]+)\s*cm', text)
        can_nang = re.search(r'Cân nặng\s*:\s*([\d.]+)\s*kg', text)
        bmi = re.search(r'BMI\s*:\s*([\d.]+)', text)
        mach = re.search(r'Mạch\s*:\s*([\d.]+)\s*l/p', text)
        huyet_ap = re.search(r'Huyết áp\s*:\s*([\d/]+)', text)

        return pd.Series([
            chieu_cao.group(1) if chieu_cao else None,
            can_nang.group(1) if can_nang else None,
            bmi.group(1) if bmi else None,
            mach.group(1) if mach else None,
            huyet_ap.group(1) if huyet_ap else None
        ], index=['Chiều cao', 'Cân nặng', 'BMI', 'Mạch', 'Huyết áp'])

    # Tách dữ liệu thể lực
    df_theluc = df_kq['Thể lực'].apply(extract_data_from_theluc)

    # Xóa cột "Thể lực"
    df_kq.drop(columns=['Thể lực'], inplace=True)

    # Chèn cột sau "Giới tính"
    cols = list(df_kq.columns)
    insert_at = cols.index('Giới tính') + 1
    for i, col_name in enumerate(['Chiều cao', 'Cân nặng', 'BMI', 'Mạch', 'Huyết áp']):
        df_kq.insert(loc=insert_at + i, column=col_name, value=df_theluc[col_name])


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Da liễu" và "Ngoại" ===========
    def classify_binh_thuong_or_bat_thuong(text):
        if text is None or pd.isna(text):
            return ""
    
        cleaned_text = str(text).strip().lower()
        
        # Nếu sau khi làm sạch mà rỗng hoặc là "nan", "none", "không khám", thì coi như không có thông tin
        if cleaned_text in ["", "nan", "none", "không khám", "chưa khám"]:
            return ""

        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
            "chưa ghi nhận bệnh lý"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        if any(keyword in cleaned_text for keyword in normal_keywords):
            return "BT"  # "BT" cho Bình thường
        else:
            return "TD"
        

    # Áp dụng hàm phân loại cho cột "Da liễu" và "Ngoại" và thay thế dữ liệu trong các cột đó
    df_kq['Da liễu'] = df_kq['Da liễu'].apply(classify_binh_thuong_or_bat_thuong)
    df_kq['Ngoại'] = df_kq['Ngoại'].apply(classify_binh_thuong_or_bat_thuong)

    def tong_hop_ngoai_khoa_da_lieu(row):
        da_lieu = row['Da liễu']
        ngoai = row['Ngoại']

        if da_lieu == 'BT' and ngoai == 'BT':
            return 'BT'
        elif da_lieu == 'TD' or ngoai == 'TD':
            return 'TD'
        else:
            return ''  # Không có dữ liệu

    df_kq['Ngoại khoa, da liễu'] = df_kq.apply(tong_hop_ngoai_khoa_da_lieu, axis=1)


    # Xóa 2 cột "Da liễu" và "Ngoại" cũ
    df_kq.drop(columns=['Da liễu', 'Ngoại'], inplace=True)

    # Di chuyển cột "Ngoại khoa, da liễu" ngay sau "Nội khoa"
    cols = df_kq.columns.tolist()
    cols.remove('Ngoại khoa, da liễu')
    noi_khoa_index = cols.index('Nội khoa') + 1
    cols.insert(noi_khoa_index, 'Ngoại khoa, da liễu')
    df_kq = df_kq[cols]


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột 'Tai mũi họng' ===========
    def classify_tai_mui_hong(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""
        
        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT" 
        
        # Danh sách từ khóa "Bệnh lý" (chỉ ví dụ, có thể mở rộng thêm)
        abnormal_keywords = [
            "viêm", 
            "nhiễm", 
            "bệnh lý", 
            "đau", 
            "chảy mủ", 
            "sưng"
        ]
        
        # Kiểm tra nếu có chứa ít nhất một từ khóa bất thường
        if any(keyword in text for keyword in abnormal_keywords):
            return "TD"  # "TD" cho Bất thường
        else:   
            return "BT"  # Nếu không có gì bất thường, mặc định là "BT"

    # Áp dụng hàm phân loại cho cột "Tai mũi họng" và thay thế dữ liệu trong cột đó
    df_kq['Tai mũi họng'] = df_kq['Tai mũi họng'].apply(classify_tai_mui_hong)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột 'Răng hàm mặt' ===========
    def classify_rang_ham_mat(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""
        
        # Danh sách từ khóa "Bình thường"
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "viêm nướu mãn", 
            "mòn răng"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT"  # "BT" cho Bình thường
        
        # Danh sách từ khóa "Bệnh lý" (chỉ ví dụ, có thể mở rộng thêm)
        abnormal_keywords = [
            "viêm", 
            "nhiễm", 
            "bệnh lý", 
            "đau", 
            "sưng", 
            "lở", 
            "chảy máu", 
            "sâu"
        ]
        
        # Kiểm tra nếu có chứa ít nhất một từ khóa bất thường
        if any(keyword in text for keyword in abnormal_keywords):
            return "TD" 
        else:
            return "BT" 

    # Áp dụng hàm phân loại cho cột "Răng hàm mặt" và thay thế dữ liệu trong cột đó
    df_kq['Răng hàm mặt'] = df_kq['Răng hàm mặt'].apply(classify_rang_ham_mat)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Sản phụ khoa" ===========
    def classify_phu_khoa(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Danh sách từ khóa "Bình thường" cho Sản phụ khoa
        normal_keywords = [
            "chưa thấy bất thường", 
            "chưa ghi nhận bất thường", 
            "không thấy bất thường", 
            "chưa ghi nhận bệnh lý"
        ]
        
        # Kiểm tra nếu có chứa bất kỳ từ khóa bình thường nào
        text = text.lower()
        if any(keyword in text for keyword in normal_keywords):
            return "BT"
        else:
            return "TD"

    # Áp dụng hàm phân loại cho cột "Sản phụ khoa" và thay thế trực tiếp dữ liệu trong cột đó
    df_kq['Sản phụ khoa'] = df_kq['Sản phụ khoa'].apply(classify_phu_khoa)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Nội khoa" ===========
    def classify_noikhoa(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Các tiểu mục bình thường của nội khoa
        normal_keywords = [
            "chưa thấy bất thường", "chưa ghi nhận bất thường", "chưa ghi nhận bệnh lý"
        ]

        # Danh sách tiểu mục kiểm tra
        tiemuc_keywords = [
            "tim mạch", "hô hấp", "tiêu hóa", "thận", "thần kinh", "nội tiết"
        ]

        # Kiểm tra nếu tất cả các tiểu mục đều bình thường (có chứa ít nhất một trong các từ khóa "bình thường")
        all_normal = all(any(normal in text.lower() for normal in normal_keywords) for keyword in tiemuc_keywords)

        if all_normal:
            return "BT"  # "BT" cho Bình thường
        else:
            return "TD"

    # Áp dụng hàm phân loại cho cột "Nội khoa" và thay thế trực tiếp dữ liệu trong cột đó
    df_kq['Nội khoa'] = df_kq['Nội khoa'].apply(classify_noikhoa)


    # ======== Xác định kết luận "Bình thường" hay "Bất thường" cho cột "Mắt" ===========
    def classify_mat(text):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ""

        # Danh sách từ khóa xác định là bình thường
        normal_keywords = [
            "chưa thấy bất thường",
            "chưa ghi nhận bệnh lý",
            "không ghi nhận bất thường"
        ]

        text_lower = text.lower()
        for keyword in normal_keywords:
            if keyword in text_lower:
                return "BT"
        
        return "TD"

    # Áp dụng vào cột 'Mắt'
    df_kq['Mắt'] = df_kq['Mắt'].apply(classify_mat)


    # Lấy dữ liệu từ cột 32 (AG), loại bỏ đầu dòng là "-" hoặc "--" và strip() hai đầu
    def clean_issues(text):
        if pd.isna(text):
            return ""
        lines = str(text).splitlines()
        cleaned = []
        for line in lines:
            line = line.strip()
            # Nếu dòng bắt đầu bằng bất kỳ ký hiệu liệt kê nào, thay thế thành '-'
            # Bao gồm: --, -, +, *, •, v.v.
            line = re.sub(r'^[\-\+\*•]{1,}', '-', line)
            # Nếu sau dấu - có khoảng trắng thì giữ lại, nếu không thì thêm khoảng trắng
            line = re.sub(r'^-(\S)', r'- \1', line)
            cleaned.append(line)
        # Loại các dòng rỗng nếu có
        cleaned = [l for l in cleaned if l]
        return '\n'.join(cleaned)
    
    # Áp dụng hàm làm sạch cho cột 30 (AE)
    # df_kq['Các vấn đề cần lưu ý'] = df.iloc[:, 30].apply(clean_issues)

    # Áp dụng hàm làm sạch cho cột 32 (AG)
    df_kq['Các vấn đề cần lưu ý'] = df.iloc[:, 32].apply(clean_issues)


    # ======= Phân loại sức khỏe =======
    # # Cột Y (24), Z (25), AA (26), AB (27), AC (28)
    # phanloai_map = {
    #     24: 'Loại I',
    #     25: 'Loại II',
    #     26: 'Loại III',
    #     27: 'Loại IV',
    #     28: 'Loại V',
    # }

    # Cột AA (26), AB (27), AC (28), AD (29), AE (30)
    phanloai_map = {
        26: 'Loại I',
        27: 'Loại II',
        28: 'Loại III',
        29: 'Loại IV',
        30: 'Loại V',
    }

    def get_phanloai(row):
        for col_idx, label in phanloai_map.items():
            val = str(row[col_idx]).strip().lower()
            if val == 'x':
                return label
        return None

    df_kq['Phân loại sức khỏe'] = df.apply(get_phanloai, axis=1)
    
    # Lưu lại vào pickle
    # save_df(df_kq, "clinical_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df_kq, out_pickle)


def run_py3(imaging_list_file, out_pickle):
    # Lấy năm hiện tại
    current_year = datetime.now().year

    # Đọc file, bỏ qua 16 dòng đầu
    df_raw = pd.read_excel(imaging_list_file, header=None, skiprows=16)
    # Chuẩn bị danh sách dữ liệu hợp lệ
    data = []

    for idx, row in df_raw.iterrows():
        # Kiểm tra Mã BA hợp lệ
        ma_ba = str(row[3]) if pd.notna(row[3]) else ""
        if not re.fullmatch(r"BA\d{8}", ma_ba):
            continue

        # Lấy họ tên từ cột G-H-I
        ho_ten = row[6] if pd.notna(row[6]) else row[7] if pd.notna(row[7]) else row[8]

        # Tuổi + Giới tính -> chuyển sang Năm sinh
        tuoi_nam = row[9]
        tuoi_nu = row[10]
        gioi_tinh = ""
        nam_sinh = ""

        if pd.notna(tuoi_nam):
            gioi_tinh = "Nam"
            nam_sinh = current_year - int(tuoi_nam)
        elif pd.notna(tuoi_nu):
            gioi_tinh = "Nữ"
            nam_sinh = current_year - int(tuoi_nu)

        # Yêu cầu và kết quả (cột S <18> và U <20>)
        yeu_cau = row[18] if pd.notna(row[18]) else row[19]
        ket_qua = row[20]

        # Nếu có đủ họ tên và yêu cầu thì thêm vào danh sách
        if pd.notna(ho_ten) and pd.notna(yeu_cau):
            data.append({
                "Mã BA": ma_ba,
                "Họ tên người bệnh": ho_ten,
                "Giới tính": gioi_tinh,
                "Năm sinh": nam_sinh,
                "Yêu cầu": yeu_cau,
                "Kết quả": ket_qua
            })

    # Tạo DataFrame từ danh sách đã xử lý
    df = pd.DataFrame(data)

    # Lưu lại vào pickle
    # save_df(df, "imaging_temp.pkl")
    # ======= Lưu ra pickle dạng file-like object =======
    pickle.dump(df, out_pickle)


def run_py4(clinical_temp_file, test_temp_file, out_pickle):
    """
    clinical_temp_file, test_temp_file: file-like object, đã mở ở chế độ 'rb'
    """
    unexpected_data_list = []
    # Load từ pickle
    df1 = pickle.load(clinical_temp_file)
    df2 = pickle.load(test_temp_file)

    def normalize_string(s):
        return str(s).strip().upper()

    df1['Họ tên chuẩn hóa'] = df1['Họ tên'].apply(normalize_string)
    df1['Giới tính chuẩn hóa'] = df1['Giới tính'].apply(normalize_string)
    df1['Năm sinh'] = df1['Năm sinh'].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")

    df2['Họ tên chuẩn hóa'] = df2['Họ tên người bệnh'].apply(normalize_string)
    df2['Giới tính chuẩn hóa'] = df2['Giới tính'].apply(normalize_string)
    df2['Năm sinh'] = df2['Năm sinh'].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")

    ds_yeu_cau = [
        'WBC', 'NEU %', 'RBC', 'HGB', 'MCV', 'MCH', 'PLT',
        'Định lượng Ure', 'Định lượng Creatinin', 'eGFR (CKD-EPI)',
        'Protein', 'Leukocytes', 'Nitrites', 'Blood',
        'Định lượng Glucose',
        'Đo hoạt độ AST (GOT)', 'Đo hoạt độ ALT (GPT)', 'Đo hoạt độ GGT',
        'Định lượng Triglycerid', 'Định lượng HDL-C', 'Định lượng LDL-C',
        'Định lượng Acid Uric', 'Bạch cầu', 'Nấm', 'Trichomonas Vaginalis'
    ]

    # df2_filtered = df2[df2['Yêu cầu'].isin(ds_yeu_cau)]

    def safe_value(x):
        allowed_strings = {'Negative', 'Positive', '1+', '2+', '3+', '4+', '+', '++', '+++', '++++'}
        if pd.isna(x):
            return ""

        # Nếu là chuỗi dạng số có dấu nháy (ví dụ: "'123.4")
        if isinstance(x, str):
            x = x.strip().lstrip("'")  # bỏ leading space và dấu nháy '
            
            # Nếu sau xử lý mà chuỗi là số thì convert
            if re.fullmatch(r'^-?\d+(\.\d+)?$', x):
                return float(x)

            # Nếu thuộc allowed_strings
            if x.upper() in allowed_strings:
                return x.upper()

            return ''

        # Nếu là số thực sự
        try:
            float(x)
            return x
        except:
            return ''
    df2_filtered = df2[df2['Yêu cầu'].isin(ds_yeu_cau)].copy()

    pivot_df = df2_filtered.pivot_table(
        index=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa', 'Năm sinh'],
        columns='Yêu cầu',
        values='Kết quả',
        aggfunc='first'
    ).reset_index()

    merged_df = pd.merge(
        df1,
        pivot_df,
        how='left',
        on=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa', 'Năm sinh']
    )

    def safe_float(x):
        try:
            if pd.isna(x) or x is None:
                return None
            if isinstance(x, str):
                x = x.strip().lstrip("'")  # bỏ dấu nháy và khoảng trắng
                if x.upper() in ['Negative', 'Positive', '1+', '2+', '3+', '4+', '5+', '+', '++', '+++', '++++', '+++++', '+-']:
                    return None
            return float(x) if x else None
        except:
            return None


    def soi_tuoi_huyet_trang(row):
        gender = str(row.get('Giới tính', '')).strip().upper()
        pk = str(row.get('Sản phụ khoa', '')).strip().upper()
        name = row.get("Họ tên", "Unknown")

        local_unexpected = []

        # Lấy dữ liệu chỉ số thô
        bc_raw = row.get('Bạch cầu', None)
        nam_raw = row.get('Nấm', None)
        tv_raw = row.get('Trichomonas Vaginalis', None)

        def is_empty(v):
            return v is None or pd.isna(v) or str(v).strip() == ""

        # === Nếu là NAM ===
        if 'NAM' in gender:
            # Nam không soi tươi: nếu hoàn toàn không có dữ liệu thì bỏ trống
            if all(is_empty(v) for v in [bc_raw, nam_raw, tv_raw]):
                return ""
            else:
                # Có dữ liệu cho nam -> log để kiểm tra nhưng không kết luận
                local_unexpected.append(
                    (
                        name,
                        "Soi tươi (nam nhưng có dữ liệu)",
                        f"Bạch cầu: {bc_raw}, Nấm: {nam_raw}, TV: {tv_raw}",
                    )
                )
                unexpected_data_list.extend(local_unexpected)
                return ""

        # === Nếu là NỮ ===
        try:
            # Nếu tất cả đều trống → chưa làm xét nghiệm
            if all(is_empty(v) for v in [bc_raw, nam_raw, tv_raw]):
                return ""

            has_any_data = False  # cờ, có ít nhất một kết quả hợp lệ không

            # ----- Bạch cầu -----
            bc_abnormal = False
            if not is_empty(bc_raw):
                try:
                    bc_str = str(bc_raw).strip().upper()
                    if "-" in bc_str:
                        parts = [p for p in bc_str.split("-") if p.strip()]
                        bc_val = safe_float(parts[-1])
                    else:
                        bc_val = safe_float(bc_str)

                    if bc_val is not None:
                        has_any_data = True
                        if bc_val > 20:
                            bc_abnormal = True
                    else:
                        local_unexpected.append((name, "Bạch cầu", bc_raw))
                except Exception:
                    local_unexpected.append((name, "Bạch cầu", bc_raw))

            # ----- Nấm -----
            nam_abnormal_values = ['1+', '2+', '3+', '4+', '+', '++', '+++', '++++']
            nam_normal_values = ['NEGATIVE']
            nam_abnormal = False

            if not is_empty(nam_raw):
                nam_str = str(nam_raw).strip().upper()
                if nam_str in nam_abnormal_values:
                    nam_abnormal = True
                    has_any_data = True
                elif nam_str in nam_normal_values:
                    has_any_data = True
                else:
                    local_unexpected.append((name, "Nấm", nam_str))

            # ----- Trichomonas Vaginalis -----
            tv_abnormal = False
            if not is_empty(tv_raw):
                tv_str = str(tv_raw).strip().upper()
                if tv_str in ['POSITIVE', 'NEGATIVE']:
                    has_any_data = True
                    if tv_str == 'POSITIVE':
                        tv_abnormal = True
                else:
                    local_unexpected.append((name, "Trichomonas Vaginalis", tv_str))

            # Ghi lại các dữ liệu không chuẩn
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Nếu thực ra không đọc được kết quả hợp lệ nào → coi như chưa làm
            if not has_any_data:
                return ""

            # Kết luận
            if bc_abnormal or nam_abnormal or tv_abnormal:
                return "TD"

            # Nếu Sản phụ khoa đã kết luận TD thì ưu tiên TD
            if pk == "TD":
                return "TD"

            return "BT"

        except Exception as e:
            unexpected_data_list.append((name, "Soi tươi", f"Exception: {e}"))
            return ""


    def tong_phan_tich(row):
        keys = ['WBC', 'NEU %', 'RBC', 'MCV', 'PLT']
        local_unexpected = []
        name = row.get("Họ tên", "Unknown")

        try:
            values = {}
            has_any_data = False  # có ít nhất 1 giá trị số hợp lệ không

            for k in keys:
                raw_val = row.get(k, None)

                # Trống / None / NaN -> không xét
                if raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "":
                    values[k] = None
                    continue

                # Convert an toàn
                val = safe_float(raw_val)
                if val is None:
                    local_unexpected.append((name, k, f"Không phải số: {raw_val}"))
                    values[k] = None
                else:
                    values[k] = val
                    has_any_data = True

            # Ghi log dữ liệu bất thường
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Không có dữ liệu số nào -> để trống
            if not has_any_data:
                return ""

            # Phân tích giá trị
            wbc = values.get('WBC')
            neu = values.get('NEU %')
            rbc = values.get('RBC')
            mcv = values.get('MCV')
            plt = values.get('PLT')

            # Nếu có bất kỳ chỉ số nào bất thường -> TD
            # WBC: 2 < WBC < 14
            if wbc is not None and not (2 < wbc < 14):
                return "TD"

            # NEU %: < 90
            if neu is not None and neu >= 90:
                return "TD"

            # RBC: 3.5 < RBC < 6
            if rbc is not None and not (3.5 < rbc < 6):
                return "TD"

            # MCV: > 80
            if mcv is not None and mcv <= 80:
                return "TD"

            # PLT: 100 < PLT < 450
            if plt is not None and not (100 < plt < 450):
                return "TD"

            # Nếu có ít nhất 1 giá trị số hợp lệ và không chỉ số nào bất thường -> BT
            return "BT"

        except Exception as e:
            unexpected_data_list.append((name, "Tổng phân tích", f"Exception: {e}"))
            return ""


    def men_gan(row):
        keys = ['Đo hoạt độ AST (GOT)', 'Đo hoạt độ ALT (GPT)']
        name = row.get("Họ tên", "Unknown")
        local_unexpected = []

        try:
            ast_raw = row.get(keys[0], None)
            alt_raw = row.get(keys[1], None)

            # Hàm parse chung cho AST/ALT
            def parse_val(raw, key_label):
                # Trống / None / NaN → coi như chưa có kết quả
                if raw is None or pd.isna(raw) or str(raw).strip() == "":
                    # Nếu muốn log “chưa có kết quả” thì mở comment:
                    # local_unexpected.append((name, key_label, "Dữ liệu trống (chưa có kết quả)"))
                    return None

                val = safe_float(raw)
                if val is None:
                    local_unexpected.append((name, key_label, f"Không phải số: {raw}"))
                return val

            ast = parse_val(ast_raw, keys[0])
            alt = parse_val(alt_raw, keys[1])

            # Ghi log bất thường nếu có
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # Không có bất kỳ kết quả nào → để trống
            if ast is None and alt is None:
                return ""

            # Nếu có bất kỳ men gan nào ≥ 90 → TD
            if (ast is not None and ast >= 90) or (alt is not None and alt >= 90):
                return "TD"

            # Có ít nhất một kết quả, tất cả đều < 90 → BT
            return "BT"

        except Exception as e:
            unexpected_data_list.append((name, "Men gan", f"Exception: {e}"))
            return ""


    def chuc_nang_than(row):
        key = 'eGFR (CKD-EPI)'
        name = row.get("Họ tên", "Unknown")
        local_unexpected = []

        try:
            raw_val = row.get(key, None)

            # 1) Trường hợp trống / chưa trả kết quả
            if raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "":
                local_unexpected.append((name, key, "Dữ liệu trống (chưa có kết quả)"))
                if local_unexpected:
                    unexpected_data_list.extend(local_unexpected)
                return ""

            # 2) Convert sang số an toàn
            val = safe_float(raw_val)
            if val is None:
                local_unexpected.append((name, key, f"Không phải số: {raw_val}"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            # 3) Đánh giá eGFR
            return "BT" if val > 80 else "TD"

        except Exception as e:
            unexpected_data_list.append((name, key, f"Exception: {e}"))
            return ""


    def duong_huyet(row):
        key = 'Định lượng Glucose'
        name = row.get("Họ tên", "Unknown")
        local_unexpected = []

        try:
            raw_val = row.get(key, None)

            # 1) Không có kết quả (trống / NaN / None) -> bỏ qua
            if raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "":
                local_unexpected.append((name, key, "Dữ liệu trống (chưa có kết quả)"))
                if local_unexpected:
                    unexpected_data_list.extend(local_unexpected)
                return ""

            # 2) Convert sang số an toàn
            val = safe_float(raw_val)
            if val is None:
                local_unexpected.append((name, key, f"Không phải số: {raw_val}"))
                unexpected_data_list.extend(local_unexpected)
                return ""

            # 3) Đánh giá BT / TD
            return "BT" if 50 <= val <= 100 else "TD"

        except Exception as e:
            unexpected_data_list.append((name, key, f"Exception: {e}"))
            return ""


    def lipid(row):
        keys = [
            'Định lượng Triglycerid',
            'Định lượng LDL-C',
            'Định lượng HDL-C',
            'Định lượng Cholesterol',
        ]
        local_unexpected = []
        results = {}  # Lưu trạng thái từng chỉ số: 'TD' hoặc 'BT'
        name = row.get("Họ tên", "Unknown")

        try:
            for key in keys:
                raw_val = row.get(key, None)

                # 1) Không có kết quả (trống / NaN / None) -> bỏ qua
                if raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "":
                    # Nếu muốn log thì uncomment:
                    # local_unexpected.append((name, key, "Dữ liệu trống (chưa có kết quả)"))
                    continue

                # 2) Convert sang số
                val = safe_float(raw_val)
                if val is None:
                    local_unexpected.append((name, key, f"Không phải số: {raw_val}"))
                    continue

                # 3) Xét kết quả cho từng chỉ số
                if key == 'Định lượng Triglycerid':
                    results[key] = 'TD' if val >= 150 else 'BT'
                elif key == 'Định lượng LDL-C':
                    results[key] = 'TD' if val >= 129 else 'BT'
                elif key == 'Định lượng HDL-C':
                    results[key] = 'TD' if val <= 40 else 'BT'
                elif key == 'Định lượng Cholesterol':
                    results[key] = 'TD' if val > 200 else 'BT'

            # Ghi log dữ liệu bất thường
            if local_unexpected:
                unexpected_data_list.extend(local_unexpected)

            # 4) Tổng hợp kết luận
            if not results:
                # Không có chỉ số nào có kết quả -> ô trống
                return ""

            status_list = list(results.values())

            if 'TD' in status_list:
                return 'TD'
            if 'BT' in status_list:
                return 'BT'

            # Fallback, gần như không xảy ra
            return ""

        except Exception as e:
            unexpected_data_list.append((name, "Exception", str(e)))
            return ""


    def sinh_hoa_khac(row):
        key = 'Định lượng Acid Uric'
        name = row.get("Họ tên", "Unknown")
        gender = str(row.get('Giới tính', '')).strip().upper()

        try:
            raw_val = row.get(key, "")

            # 1) Nếu trống, None, hoặc NaN → không kết luận
            if raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "":
                return ""

            # 2) Kiểm tra có phải số hay không
            val = safe_float(raw_val)
            if val is None:  # safe_float không convert được
                unexpected_data_list.append((name, key, f"Không phải số: {raw_val}"))
                return ""

            # 3) Xử lý logic BT/TD
            if ('NAM' in gender and val >= 420) or ('NỮ' in gender and val >= 360):
                return "TD"
            elif 'NAM' in gender and val < 420:
                return 'BT'
            elif 'NỮ' in gender and val < 360:
                return 'BT'

            # 4) Trường hợp giới tính lạ
            unexpected_data_list.append((name, "Giới tính", gender))
            return ""

        except Exception as e:
            unexpected_data_list.append((name, "Exception", str(e)))
            return ""


    def nuoc_tieu(row):
        keys = ['Protein', 'Leukocytes', 'Nitrites', 'Blood']
        abnormal_values = ['POSITIVE', '2+', '3+', '4+', '5+', '6+', '7+', '++', '+++', '++++', '+++++', '++++++', '+++++++']
        normal_values = ['NEGATIVE', '1+', '0', '+-']
        unexpected_data = []
        values = []
        abnormal_found = False
        empty_count = 0  # Đếm số ô trống

        try:
            for k in keys:
                val = row.get(k, None)
                val_str = str(val).strip().upper() if val is not None else ""

                if val_str == "" or val_str == "NAN":
                    empty_count += 1
                    continue

                if k in ['Leukocytes', 'Blood']:
                    try:
                        num_val = float(val_str)
                        if k == 'Blood' and num_val >= 80:
                            return "TD"
                        elif k == 'Leukocytes' and num_val > 100:
                            return "TD"
                        else:
                            values.append("BT")
                            continue
                    except Exception:
                        pass  # không phải số, xử lý tiếp như chuỗi

                if val_str in abnormal_values:
                    abnormal_found = True
                elif val_str in normal_values:
                    values.append("BT")
                else:
                    unexpected_data.append((row.get('Họ tên', 'Không rõ'), k, val))
            
            if unexpected_data:
                unexpected_data_list.extend(unexpected_data)

            if abnormal_found:
                return "TD"

            if empty_count == len(keys):
                return ""

            if values and all(v == "BT" for v in values):
                return "BT"

            return ""  # nếu không rõ ràng thì để trống

        except Exception as e:
            unexpected_data_list.append((row.get('Họ tên', 'Không rõ'), 'nuoc_tieu', str(e)))
            return str(e)


        
    merged_df['Soi tươi huyết trắng'] = merged_df.apply(soi_tuoi_huyet_trang, axis=1)
    merged_df['Tổng phân tích tế bào máu ngoại vi'] = merged_df.apply(tong_phan_tich, axis=1)
    merged_df['Men gan'] = merged_df.apply(men_gan, axis=1)
    merged_df['Chức năng thận'] = merged_df.apply(chuc_nang_than, axis=1)
    merged_df['Đường huyết lúc đói'] = merged_df.apply(duong_huyet, axis=1)
    merged_df['Xét nghiệm lipid máu'] = merged_df.apply(lipid, axis=1)
    merged_df['Sinh hóa khác'] = merged_df.apply(sinh_hoa_khac, axis=1)
    merged_df['Tổng phân tích nước tiểu'] = merged_df.apply(nuoc_tieu, axis=1)

    merged_df.drop(columns=[col for col in ds_yeu_cau if col in merged_df.columns], inplace=True)
    merged_df.drop(columns=['Họ tên chuẩn hóa', 'Giới tính chuẩn hóa'], inplace=True)

    pickle.dump(merged_df, out_pickle)
    return unexpected_data_list


def run_full_pipeline(customer_list, test_list, clinical_list, imaging_list, output_path):
    """
    Các biến customer_list, test_list,... là file-like object (request.FILES['...']).
    output_path: path hoặc file-like object để xuất báo cáo.
    """
    # ==== Bước 1: Xử lý test_list, clinical_list, imaging_list ====
    # Tạo buffer memory để lưu pickle tạm
    test_temp = io.BytesIO()
    clinical_temp = io.BytesIO()
    imaging_temp = io.BytesIO()
    merge_clinical_test_temp = io.BytesIO()

    # run_py1, run_py2, run_py3 nhận file-like object (Excel),
    # lưu kết quả ra buffer pickle
    run_py1(test_list, out_pickle=test_temp)
    run_py2(clinical_list, out_pickle=clinical_temp)
    run_py3(imaging_list, out_pickle=imaging_temp)

    # Reset pointer để đọc lại pickle
    test_temp.seek(0)
    clinical_temp.seek(0)
    imaging_temp.seek(0)

    # ==== Bước 2: So sánh bất thường ====
    # run_py4 trả về pickle cho bước sau:
    merge_clinical_test_temp = io.BytesIO()
    unexpected_data_all = []
    unexpected_data_all.append(
        run_py4(clinical_temp, test_temp, out_pickle=merge_clinical_test_temp)
    )
    merge_clinical_test_temp.seek(0)

    # ==== Bước 3: Ghép dữ liệu & xuất báo cáo ====
    # run_py5 nhận pickle, excel file-like, trả file excel
    run_py5(
        merged_file=merge_clinical_test_temp,  
        clinical_list_file=clinical_list,
        imaging_temp_file=imaging_temp,
        customer_list_file=customer_list,
        output_file=output_path
    )

    return unexpected_data_all

