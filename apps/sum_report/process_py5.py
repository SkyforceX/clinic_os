import pickle
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl import load_workbook
from datetime import date
import re
from django.utils.text import slugify
import io

def run_py5(
    merged_file, clinical_list_file, imaging_temp_file, customer_list_file, output_file
):
    # === 1. Load dữ liệu pickle và excel từ file-like object ===
    df_base = pickle.load(merged_file)
    df_cdha = pickle.load(imaging_temp_file)

    # Chuẩn hóa tên
    df_base['Họ tên'] = df_base['Họ tên'].astype(str).str.strip()
    df_base['Giới tính'] = df_base['Giới tính'].astype(str).str.strip()
    df_base['Năm sinh'] = df_base['Năm sinh'].astype(str).str.strip()
    df_cdha['Họ tên'] = df_cdha['Họ tên người bệnh'].astype(str).str.strip()
    df_cdha['Giới tính'] = df_cdha['Giới tính'].astype(str).str.strip()
    df_cdha['Năm sinh'] = df_cdha['Năm sinh'].astype(str).str.strip()

    # === 2. Lọc & pivot các dịch vụ chẩn đoán hình ảnh ===
    dich_vu_chon = [
        'Chụp X-quang Ngực thẳng',
        'Siêu âm tuyến giáp',
        'Siêu âm tuyến vú hai bên'
    ]
    df_cdha_filtered = df_cdha[
        df_cdha['Yêu cầu'].isin(dich_vu_chon) |
        df_cdha['Yêu cầu'].str.contains('Siêu âm ổ bụng', na=False)
    ]
    df_cdha_filtered.loc[
        df_cdha_filtered['Yêu cầu'].str.contains('Siêu âm ổ bụng', na=False),
        'Yêu cầu'
    ] = 'Siêu âm ổ bụng'

    df_pivot = df_cdha_filtered.pivot_table(
        index=['Họ tên', 'Giới tính', 'Năm sinh'],
        columns='Yêu cầu',
        values='Kết quả',
        aggfunc='first'
    ).reset_index()

    df_merged = pd.merge(df_base, df_pivot, on=['Họ tên', 'Giới tính', 'Năm sinh'], how='left')

    # Đánh giá BT/TD cho các dịch vụ
    cot_danh_gia = {
        'Chụp X-quang Ngực thẳng': ['chưa thấy bất thường', 'trong giới hạn bình thường'],
        'Siêu âm tuyến giáp': ['chưa thấy bất thường', 'chưa phát hiện bất thường'],
        'Siêu âm tuyến vú hai bên': ['chưa thấy bất thường', 'chưa phát hiện bất thường']
    }
    def danh_gia(text, dieu_kien_bt):
        if text is None or pd.isna(text) or str(text).strip().lower() in ["", "nan", "none"]:
            return ''
        val_lower = str(text).lower()
        for dk in dieu_kien_bt:
            if dk in val_lower:
                return 'BT'
        return 'TD'
    def danh_gia_sieu_am_o_bung(val):
        if pd.isna(val): return ''
        return 'BT' if 'chưa phát hiện bất thường' in str(val).lower() else 'TD'
    for col, dieu_kien_bt in cot_danh_gia.items():
        if col in df_merged.columns:
            df_merged[col] = df_merged[col].apply(lambda x: danh_gia(x, dieu_kien_bt))
    if 'Siêu âm ổ bụng' in df_merged.columns:
        df_merged['Siêu âm ổ bụng'] = df_merged['Siêu âm ổ bụng'].apply(danh_gia_sieu_am_o_bung)

    # Đảm bảo các cột cần thiết cuối bảng
    cols = [col for col in df_merged.columns if col not in ['Các vấn đề cần lưu ý', 'Phân loại sức khỏe']]
    if 'Các vấn đề cần lưu ý' in df_merged.columns: cols.append('Các vấn đề cần lưu ý')
    if 'Phân loại sức khỏe' in df_merged.columns: cols.append('Phân loại sức khỏe')
    df_merged = df_merged[cols]

    # === Bổ sung và chuẩn hóa thông tin từ DS_KSK.xlsx ===
    # Hàm chuẩn hóa mã BN, chấp nhận mọi kiểu nhập, trả về BN + 6 số cuối cùng (giống BN0001140, bn01140, 1140 đều được)
    def standardize_bn(x):
        s = str(x).strip().upper()
        m = re.search(r'(\d+)$', s)
        if m:
            num = m.group(1)
            # Chuẩn hóa về 8 số cuối cùng
            return 'BN' + num.zfill(8)
        return ''

    # Chuẩn hóa họ tên (bỏ dấu, viết thường, loại khoảng trắng)
    import unidecode
    def standardize_name(name):
        return unidecode.unidecode(str(name)).strip().lower()

    
    # Đọc file và chuẩn hóa tên cột
    df_ksk = pd.read_excel(customer_list_file, skiprows=1, dtype=str)
    df_ksk.columns = df_ksk.columns.str.strip().str.lower()
    # pattern = re.compile(r'^BN\d{6}$')
    # df_filtered = df_ksk[df_ksk['mã bn'].apply(lambda x: bool(pattern.match(str(x).strip())))].copy()
    # df_filtered['mã bn'] = df_filtered['mã bn'].str.strip()
    # df_filtered['họ tên'] = df_filtered['họ tên'].str.strip().str.lower()
    # df_filtered['giới tính'] = df_filtered['giới tính'].str.strip().str.capitalize()
    # df_filtered['ngày sinh'] = pd.to_datetime(df_filtered['ngày sinh'], errors='coerce')
    # df_filtered['ngày sinh dạng chuỗi'] = df_filtered['ngày sinh'].dt.strftime('%d/%m/%Y')
    # df_filtered['năm sinh'] = df_filtered['ngày sinh'].dt.year.astype('Int64')
    # ksk_dict = {
    #     (row['họ tên'], row['mã bn']): (row['ngày sinh dạng chuỗi'], row['năm sinh'], row['giới tính'])
    #     for _, row in df_filtered.iterrows()
    # }

    # 1 start. Áp dụng chuẩn hóa
#     df_ksk['mã bn'] = df_ksk['mã bn'].apply(standardize_bn)
#     df_ksk['giới tính'] = df_ksk['giới tính'].str.strip().str.capitalize()
#     df_ksk['ngày sinh'] = pd.to_datetime(df_ksk['ngày sinh'], errors='coerce')
#     df_ksk['ngày sinh'] = df_ksk['ngày sinh'].dt.strftime('%d/%m/%Y')
    
#     # Chỉ lấy mã BN đúng chuẩn (6 số trở lên, cho an toàn)
#     df_filtered = df_ksk[df_ksk['mã bn'].str.upper().str.match(r'^BN\d{6,}$', na=False)].copy()
#     df_filtered['key'] = df_filtered['mã bn']

#    # Chuẩn hóa mã BN ở bảng kết quả (df_merged):
#     if 'Mã bệnh nhân' in df_merged.columns:
#         df_merged['Mã bệnh nhân'] = df_merged['Mã bệnh nhân'].apply(standardize_bn)
#         df_merged['key'] = df_merged['Mã bệnh nhân']
#     else:
#         df_merged['key'] = df_merged['Họ tên']  # fallback, không nên dùng

#     # Merge giữ nguyên họ tên bên danh sách khách hàng
#     df_out = pd.merge(
#         df_filtered,
#         df_merged.drop_duplicates('key'),  # loại trùng nếu có
#         on='key',
#         how='left',
#         suffixes=('', '_y')
#     )

#     # Danh sách 4 cột giữ từ DS_KSK (sau chuẩn hóa/tạo lại)
#     main_cols = ['họ tên', 'mã bn', 'giới tính', 'ngày sinh']
#     df_out = df_out[main_cols + [col for col in df_out.columns if col not in main_cols + ['key']]]

#     # Danh sách các cột từ df_merged (cột kết quả)  
#     # Loại bỏ các cột trùng tên (kể cả dạng hoa/thường, dạng có _y, ...)
#     cols_result = [col for col in df_merged.columns if col.lower() not in ['họ tên', 'mã bệnh nhân', 'giới tính', 'ngày sinh']]

#     # 4. Ghép lại bảng theo thứ tự: 4 cột DS_KSK + các cột kết quả
#     df_out = df_out[['họ tên', 'mã bn', 'giới tính', 'ngày sinh'] + cols_result]

#     # 5. Nếu cần STT:
#     df_out.insert(0, 'STT', range(1, len(df_out) + 1))

#     # 6 end. Xuất file
#     df_out.to_excel(output_file, index=False)

    # 1. Áp dụng chuẩn hóa
    df_ksk['mã bn'] = df_ksk['mã bn'].apply(standardize_bn)
    df_ksk['giới tính'] = df_ksk['giới tính'].str.strip().str.capitalize()

    # xử lý chuẩn hóa ngày sinh
    def clean_date(x):
        if pd.isna(x):
            return ""
        x = str(x).strip()

        # TH Excel lưu dạng số serial
        if x.replace('.', '', 1).isdigit():
            try:
                return pd.to_datetime(float(x), unit='d', origin='1899-12-30').strftime('%d/%m/%Y')
            except:
                pass

        # Thử các pattern phổ biến
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return pd.to_datetime(x, format=fmt).strftime("%d/%m/%Y")
            except:
                pass

        # Cuối cùng: để pandas tự đoán
        try:
            return pd.to_datetime(x, dayfirst=True).strftime("%d/%m/%Y")
        except:
            return ""  # tránh NaT → mất dữ liệu

    df_ksk['ngày sinh'] = df_ksk['ngày sinh'].apply(clean_date)

    # 2. Chỉ lấy mã BN đúng chuẩn (6 số trở lên)
    df_filtered = df_ksk[df_ksk['mã bn'].str.upper().str.match(r'^BN\d{6,}$', na=False)].copy()
    df_filtered['key'] = df_filtered['mã bn']

    # 3. Chuẩn hóa mã BN ở bảng kết quả (df_merged)
    if 'Mã bệnh nhân' in df_merged.columns:
        df_merged['Mã bệnh nhân'] = df_merged['Mã bệnh nhân'].apply(standardize_bn)
        df_merged['key'] = df_merged['Mã bệnh nhân']
    else:
        df_merged['key'] = df_merged['Họ tên']  # fallback, không nên dùng

    # 4. Merge giữ nguyên họ tên bên DS_KSK, loại trùng theo key
    df_out = pd.merge(
        df_filtered,
        df_merged.drop_duplicates('key'),
        on='key',
        how='left',
        suffixes=('', '_y')
    )

    # 5. Danh sách 4 cột chính giữ từ DS_KSK
    main_cols = ['họ tên', 'mã bn', 'giới tính', 'ngày sinh']

    # 6. Lấy danh sách cột kết quả từ df_merged (bỏ 4 cột chính và 'key', không phân biệt hoa/thường)
    cols_result = [
        col for col in df_merged.columns
        if col.lower() not in ['họ tên', 'mã bệnh nhân', 'giới tính', 'ngày sinh', 'key']
    ]

    # 7. Chỉ lấy các cột kết quả thực sự có trong df_out (tránh lỗi thiếu cột khi merge)
    cols_exist = [col for col in cols_result if col in df_out.columns]

    # 8. Đảm bảo đúng thứ tự: 4 cột DS_KSK + các cột kết quả
    final_cols = main_cols + cols_exist

    # 9. Tạo bảng kết quả cuối cùng
    df_out = df_out[final_cols].copy()

    # 10. Thêm STT nếu muốn
    df_out.insert(0, 'STT', range(1, len(df_out) + 1))

    # 11. Xuất file
    df_out.to_excel(output_file, index=False)


    # Định dạng bảng Excel với openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # định dạng in
    # 1. Đặt khổ giấy A4 nằm ngang
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # 2. Điều chỉnh lề
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

    # 3. Lặp lại dòng tiêu đề (dòng 1 hoặc 2 tùy bạn)
    ws.print_title_rows = "1:2"  # Lặp lại 2 dòng đầu, nếu chỉ 1 dòng thì "1:1"

    # 4. Chỉnh cỡ chữ toàn bảng (tùy chọn, giúp bảng nhỏ gọn hơn)
    from openpyxl.styles import Font
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.font = Font(size=10)  # hoặc 9

    # 5. Fit bảng vào khổ in ngang A4 (scale cho vừa trang)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0   # 0 = tự động tính số trang theo chiều dọc

    # Lặp lại dòng 1 và 2 ở đầu mỗi trang (chỉnh lại nếu chỉ cần 1 dòng)
    ws.row_dimensions[1].height = 30
    ws.print_title_rows = "1:2"  
    # ---------------------------------------------------------------------

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)

    # Định dạng từng ô (tối ưu)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(text_rotation=90, vertical='center', horizontal='center', wrap_text=True)
            else:
                value = str(cell.value).strip() if cell.value is not None else ''
                col_name = ws.cell(row=1, column=cell.column).value
                cac_cot_can_giua = [
                    'STT', 'Mã bệnh nhân', 'Năm sinh', 'Giới tính', 'Chiều cao', 'Cân nặng', 'Huyết áp', 'BMI',
                    'Mạch', 'Nhiệt độ', 'Nhịp thở', 'Thị lực P', 'Thị lực T', 'Phân loại sức khỏe'
                ]
                if value in ['BT', 'TD'] or col_name in cac_cot_can_giua:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_name == 'Các vấn đề cần lưu ý':
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
    ws.row_dimensions[1].height = 70

    # Xoay dọc mã BN và năm sinh (C và D)
    row = 2
    while True:
        cell_stt = ws.cell(row=row, column=1)
        try:
            if cell_stt.value is None or not isinstance(int(cell_stt.value), int):
                break
        except:
            break
        align = Alignment(textRotation=90, vertical='center', horizontal='center')
        ws.cell(row=row, column=3).alignment = align
        ws.cell(row=row, column=4).alignment = align
        row += 1

    # === Thêm dòng tiêu đề lớn (tên công ty + tiêu đề báo cáo) ở đầu ===
    ws.insert_rows(1)
    clinical_list_file.seek(0)
    df_raw_clinical = pd.read_excel(clinical_list_file, header=None)
    title_file = df_raw_clinical.iloc[5, 2] if not df_raw_clinical.empty else "UnknownCompany"
    company_name = (
        title_file.replace("DANH SÁCH KHÁM SỨC KHỎE ĐỊNH KỲ CBCNV ", "").strip()
        if title_file else "UnknownCompany"
    )
    full_title = f"{company_name}\nKẾT QUẢ KIỂM TRA SỨC KHỎE TỔNG QUÁT NĂM {date.today().year}"
    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = full_title
    title_cell.font = Font(name="Times New Roman", size=18, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_formula_cell = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_stat_header = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # === Thống kê cuối bảng ===
    start_data_row = 3
    end_data_row = start_data_row
    while True:
        cell_stt = ws.cell(row=end_data_row, column=1)
        try:
            if cell_stt.value is None or not isinstance(int(cell_stt.value), int):
                break
        except:
            break
        end_data_row += 1
    end_data_row -= 1
    summary_titles = [
        "TỔNG SỐ THỰC HIỆN",
        "TỔNG SỐ BÌNH THƯỜNG (BT)",
        "TỔNG SỐ BẤT THƯỜNG CẦN THEO DÕI THÊM (TD)",
        "TỔNG SỐ TỪ CHỐI (TC)",
        "TỔNG SỐ KHÔNG THỰC HIỆN (Ô TRỐNG)",
        "TỈ LỆ BẤT THƯỜNG CẦN THEO DÕI THÊM",
        "TỔNG KHÁCH HÀNG"
    ]
    summary_start_row = end_data_row + 1
    summary_end_row = summary_start_row + len(summary_titles) - 1
    for i, title in enumerate(summary_titles):
        row_ = summary_start_row + i
        ws.merge_cells(start_row=row_, start_column=1, end_row=row_, end_column=10)
        cell = ws.cell(row=row_, column=1, value=title)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = fill_stat_header
        for col in range(11, max_col):
            col_letter = get_column_letter(col)
            cell2 = ws.cell(row=row_, column=col)
            data_col_range = f"{col_letter}{start_data_row}:{col_letter}{end_data_row}"
            if i == 0:
                cell2.value = f"=COUNTA({data_col_range})"
            elif i == 1:
                cell2.value = f'=COUNTIF({data_col_range},"BT")'
            elif i == 2:
                cell2.value = f'=COUNTIF({data_col_range},"TD")'
            elif i == 3:
                cell2.value = f'=COUNTIF({data_col_range},"TC")'
            elif i == 4:
                cell2.value = f'=COUNTBLANK({data_col_range})'
            elif i == 5:
                cell2.value = f'=IFERROR(COUNTIF({data_col_range},"TD")/COUNTA({data_col_range}), "")'
                cell2.number_format = '0 %'
            elif i == 6:
                cell2.value = f'=COUNTA(B{start_data_row}:B{end_data_row})'
            cell2.alignment = Alignment(horizontal="center")
            if col != 29 and col != 30:
                cell2.fill = fill_formula_cell
    # Merge ô ký, ngày tháng cuối
    ws.merge_cells(
        start_row=summary_start_row, start_column=max_col - 1,
        end_row=summary_end_row, end_column=max_col
    )
    main_cell = ws.cell(row=summary_start_row, column=max_col - 1)
    main_cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
    main_cell.font = Font(italic=True)

    # Kẻ viền cho 7 dòng cuối
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=summary_start_row, max_row=summary_end_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    # ===== Lưu workbook ===== #
    wb.save(output_file)
    wb.close()