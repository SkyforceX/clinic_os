from io import BytesIO
import re
import unicodedata

from django.db import transaction
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.contract.models import ImplementationPlan, ImplementationPlanLog


THIN = Side(style="thin", color="C9D2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="EAF3FB")


DEPARTMENT_DEFINITIONS = [
    {
        "key": "sales",
        "label": "Sales",
        "aliases": ["sales", "sale", "hc sale"],
        "group_names": {"Sales", "Sale", "HC Sale", "Marketing", "Business"},
    },
    {
        "key": "tkyk",
        "label": "TKYK",
        "aliases": ["tkyk"],
        "group_names": {"TKYK"},
    },
    {
        "key": "accounting",
        "label": "Kế toán",
        "aliases": ["ke toan", "kế toán", "ke toan", "accounting"],
        "group_names": {"Kế toán", "Ke toan", "Accounting", "Finance"},
    },
    {
        "key": "operations",
        "label": "Vận hành",
        "aliases": ["vh", "van hanh", "vận hành", "operations"],
        "group_names": {"VH", "Vận hành", "Van hanh", "Operations"},
    },
    {
        "key": "nursing",
        "label": "Điều dưỡng",
        "aliases": ["dd", "đd", "dieu duong", "điều dưỡng", "nurse", "nursing"],
        "group_names": {"Điều dưỡng", "Dieu duong", "Nurses", "Nursing", "DD", "ĐD"},
    },
    {
        "key": "doctor",
        "label": "Bác sĩ",
        "aliases": ["bs", "bac si", "bác sĩ", "doctor", "doctors"],
        "group_names": {"Bác sĩ", "Bac si", "Doctor", "Doctors", "BS"},
    },
    {
        "key": "customer_service",
        "label": "CSKH",
        "aliases": ["cskh", "cham soc khach hang", "chăm sóc khách hàng", "customer service"],
        "group_names": {"CSKH", "Chăm sóc khách hàng", "Cham soc khach hang", "Customer Service"},
    },
]
DEPARTMENT_ORDER = [item["key"] for item in DEPARTMENT_DEFINITIONS]
DEPARTMENT_MAP = {item["key"]: item for item in DEPARTMENT_DEFINITIONS}
ALL_DEPARTMENT_KEYS = list(DEPARTMENT_ORDER)


IMPLEMENTATION_TEMPLATE = [
    {
        "owner": "Các BP",
        "category": "__COMPANY_NAME__",
        "detail_type": "company_short_name",
        "detail_default": "",
        "note_default": "Để các BP nhận biết khi giao tiếp với KH",
    },
    {
        "owner": "Các BP",
        "category": "ĐỊA CHỈ",
        "detail_type": "company_address",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Kế toán",
        "category": "MST",
        "detail_type": "tax_code",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Các BP",
        "category": "NGƯỜI LIÊN HỆ",
        "detail_type": "contact_person",
        "detail_default": "",
        "note_default": "Liên hệ khi có phát sinh sự việc",
    },
    {
        "owner": "Các BP",
        "category": "SỐ LƯỢNG NHÂN VIÊN (LẤY MÁU)",
        "detail_type": "employee_count",
        "detail_default": "",
        "note_default": "Chi tiết theo Danh sách đính kèm",
    },
    {
        "owner": "VH + ĐD + Sales",
        "category": "THỜI GIAN LẤY MÁU",
        "detail_type": "blood_collection_time",
        "detail_default": "",
        "note_default": "Các ca chưa lấy máu tại VP CTY, sẽ ghé VMD lấy",
    },
    {
        "owner": "VH + ĐD + Sales",
        "category": "ĐỊA ĐIỂM LẤY MÁU",
        "detail_type": "blood_collection_location",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "Các BP",
        "category": "THỜI GIAN KHÁM",
        "detail_type": "checkup_time",
        "detail_default": "",
        "note_default": "",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "DANH MỤC KHÁM NGOÀI GÓI",
        "detail_type": "fixed",
        "detail_default": "TOÀN BỘ CHI PHÍ PHÁT SINH của Nhân viên và Người thân sẽ tự thanh toán.",
        "note_default": "Người thân sẽ do CTY thông báo",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "ƯU ĐÃI: NGƯỜI THÂN",
        "detail_type": "fixed",
        "detail_default": "Áp dụng cho người thân theo Danh sách Cty đã xác nhận",
        "note_default": "TỰ THANH TOÁN",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "ƯU ĐÃI GIẢM GIÁ (áp dụng cho nhân viên và Người thân SAU KHI đã sử dụng Gói khám CTY)",
        "detail_type": "fixed",
        "detail_default": "Giảm 10% các dịch vụ ngoài Gói CTY, ngoại trừ (Thuốc, Nội Soi, Phẫu thuật, Thủ Thuật, các dịch vụ có sử dụng Vật tư tiêu hao đặc biệt.)",
        "note_default": "TỰ THANH TOÁN",
    },
    {
        "owner": "TKYK - Kế Toán",
        "category": "PHÁT SINH CHUYỂN ĐỔI GÓI: NỮ ĐT --> <-- NỮ GĐ",
        "detail_type": "fixed",
        "detail_default": "Nhân viên ĐƯỢC / KHÔNG chuyển đổi Gói Nữ ĐT sang Gói Nữ GĐ",
        "note_default": "",
    },
    {
        "owner": "Sales - TKYK",
        "category": "CTY CÓ YÊU CẦU CUNG CẤP GIẤY CNSK TT32 ?",
        "detail_type": "fixed",
        "detail_default": "KHÔNG YÊU CẦU / CÓ YÊU CẦU",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "THỜI HẠN & HÌNH THỨC TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "Thời hạn: 10 ngày, kể từ ngày KH đã hoàn tất Gói khám\nHình thức: TRẢ 1 LẦN / CUỐN CHIẾU.",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "ĐỊA CHỈ TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "* Hồ sơ cá nhân:\n* SMR:",
        "note_default": "",
    },
    {
        "owner": "TKYK",
        "category": "NGƯỜI LIÊN HỆ TRẢ HỒ SƠ",
        "detail_type": "fixed",
        "detail_default": "Sales sẽ thông báo",
        "note_default": "",
    },
    {
        "owner": "BS - Điều Dưỡng",
        "category": "PHÂN LOẠI KHÁM SỨC KHỎE",
        "detail_type": "fixed",
        "detail_default": "Có",
        "note_default": "",
    },
    {
        "owner": "TKYK - BS - Điều Dưỡng",
        "category": "BÁO CÁO TỔNG KẾT (SUMMARY REPORT)",
        "detail_type": "fixed",
        "detail_default": "Tiếng Việt",
        "note_default": "",
    },
    {
        "owner": "BS - Điều Dưỡng",
        "category": "PHÁT HIỆN BỆNH LÝ",
        "detail_type": "fixed",
        "detail_default": "BS / Điều dưỡng thông tin cho Bệnh nhân và Sales (nếu có liên quan đến Bệnh truyền nhiễm).",
        "note_default": "",
    },
    {
        "owner": "Kế toán - Sales",
        "category": "THỜI HẠN THANH TOÁN",
        "detail_type": "fixed",
        "detail_default": "10 ngày, kể từ ngày Kế toán xuất hóa đơn tài chính",
        "note_default": "",
    },
    {
        "owner": "CSKH - Sales",
        "category": "BÁO CÁO ĐÁNH GIÁ DỊCH VỤ TỪ KHÁCH HÀNG",
        "detail_type": "fixed",
        "detail_default": "Gửi KH Form Đánh giá dịch vụ --> Ghi nhận ý kiến KH --> QL Chất lượng sau khi hoàn tất KSK",
        "note_default": "",
    },
]


def _clean_spaces(value):
    return re.sub(r"\s+", " ", (value or "").strip())


def _normalize_text(value):
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return _clean_spaces(text)


def _actor_display_name(actor):
    if not actor:
        return "Hệ thống"
    return actor.get_full_name() or actor.username or f"User #{actor.pk}"


def build_company_short_name(name: str) -> str:
    raw = _clean_spaces(name).upper()
    if not raw:
        return ""

    stop_words = {
        "CÔNG",
        "TY",
        "TNHH",
        "MTV",
        "TM",
        "DV",
        "DỊCH",
        "VỤ",
        "CP",
        "CỔ",
        "PHẦN",
        "DOANH",
        "NGHIỆP",
        "TRÁCH",
        "NHIỆM",
        "HỮU",
        "HẠN",
        "CHI",
        "NHÁNH",
        "VÀ",
    }
    words = [w for w in re.split(r"[^\wÀ-ỹ]+", raw) if w and w not in stop_words]
    if not words:
        return raw[:30]

    short_name = "".join(word[0] for word in words[:6])
    return short_name[:20]


def build_contact_display(contract) -> str:
    user = getattr(contract, "created_by", None)
    if not user:
        return ""
    return getattr(user, "get_full_name", lambda: "")() or getattr(user, "username", "") or ""


def format_date(value):
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def format_datetime(value):
    if not value:
        return ""
    return value.strftime("%H:%M, %d/%m/%Y")


def build_blood_collection_time(contract, profile):
    if profile and getattr(profile, "blood_collection_from_at", None):
        if getattr(profile, "blood_collection_to_at", None):
            return f"{format_datetime(profile.blood_collection_from_at)} - {format_datetime(profile.blood_collection_to_at)}"
        return format_datetime(profile.blood_collection_from_at)

    first_row = contract.blood_collection_schedules.order_by("collection_date", "id").first()
    if first_row and first_row.collection_date:
        return format_date(first_row.collection_date)
    return ""


def build_blood_collection_location(contract, profile):
    if profile and getattr(profile, "blood_collection_location", None):
        return profile.blood_collection_location or ""
    first_row = contract.blood_collection_schedules.order_by("collection_date", "id").first()
    return getattr(first_row, "location", "") or ""


def build_checkup_time(contract):
    if contract.start_date and contract.end_date:
        return f"{format_date(contract.start_date)} - {format_date(contract.end_date)}"
    if contract.start_date:
        return format_date(contract.start_date)
    if contract.end_date:
        return format_date(contract.end_date)
    return ""


def build_employee_count(contract, profile):
    if contract.employee_count:
        return str(contract.employee_count)

    if profile:
        total = (
            int(getattr(profile, "male_count", 0) or 0)
            + int(getattr(profile, "female_single_count", 0) or 0)
            + int(getattr(profile, "female_family_count", 0) or 0)
        )
        if total > 0:
            return str(total)

    return ""


def build_dynamic_value(contract, profile, detail_type, category):
    company = contract.company
    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif company:
        company_name = getattr(company, "name", "") or ""

    if detail_type == "company_short_name":
        short_name = build_company_short_name(company_name)
        return f"Gọi tắt: {short_name}" if short_name else ""
    if detail_type == "company_address":
        if profile and getattr(profile, "company_address_snapshot", None):
            return profile.company_address_snapshot or ""
        return getattr(company, "address", "") or ""
    if detail_type == "tax_code":
        if profile and getattr(profile, "company_tax_code_snapshot", None):
            return profile.company_tax_code_snapshot or ""
        return getattr(company, "tax_code", "") or ""
    if detail_type == "contact_person":
        return build_contact_display(contract)
    if detail_type == "employee_count":
        return build_employee_count(contract, profile)
    if detail_type == "blood_collection_time":
        return build_blood_collection_time(contract, profile)
    if detail_type == "blood_collection_location":
        return build_blood_collection_location(contract, profile)
    if detail_type == "checkup_time":
        return build_checkup_time(contract)
    return ""


def resolve_row_department_keys(owner_text):
    owner_normalized = _normalize_text(owner_text)
    if not owner_normalized:
        return []
    if "cac bp" in owner_normalized:
        return list(ALL_DEPARTMENT_KEYS)

    matches = []
    for department in DEPARTMENT_DEFINITIONS:
        if any(alias in owner_normalized for alias in department["aliases"]):
            matches.append(department["key"])
    return sorted(set(matches), key=lambda item: DEPARTMENT_ORDER.index(item))


def get_department_label(department_key):
    return DEPARTMENT_MAP.get(department_key, {}).get("label", department_key)


def get_user_department_keys(user):
    if not user or not getattr(user, "is_authenticated", False):
        return []

    group_names = set(user.groups.values_list("name", flat=True))
    keys = []
    for department in DEPARTMENT_DEFINITIONS:
        if group_names.intersection(department["group_names"]):
            keys.append(department["key"])

    if getattr(user, "is_superuser", False):
        return list(ALL_DEPARTMENT_KEYS)
    return sorted(set(keys), key=lambda item: DEPARTMENT_ORDER.index(item))


def build_default_rows(contract):
    profile = getattr(contract, "corporate_profile", None)
    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif contract.company:
        company_name = contract.company.name

    rows = []
    for idx, item in enumerate(IMPLEMENTATION_TEMPLATE, start=1):
        category = item["category"]
        if category == "__COMPANY_NAME__":
            category = company_name or "TÊN CÔNG TY"

        detail = item["detail_default"]
        if item["detail_type"] != "fixed":
            detail = build_dynamic_value(contract, profile, item["detail_type"], category)

        rows.append(
            {
                "stt": idx,
                "owner": item["owner"],
                "category": category,
                "detail": detail or "",
                "note": item["note_default"] or "",
                "department_keys": resolve_row_department_keys(item["owner"]),
                "confirmations": {},
                "is_locked": False,
                "locked_at": "",
                "locked_by_name": "",
            }
        )
    return rows


def normalize_rows(rows, contract):
    defaults = build_default_rows(contract)
    if not rows:
        return defaults

    normalized = []
    for idx, default in enumerate(defaults):
        existing = rows[idx] if idx < len(rows) else {}
        confirmations = existing.get("confirmations") or {}
        department_keys = existing.get("department_keys") or default["department_keys"]

        cleaned_confirmations = {}
        for department_key in department_keys:
            status = confirmations.get(department_key) or {}
            if not status:
                continue
            confirmed_at = status.get("confirmed_at") or ""
            confirmed_by_name = status.get("confirmed_by_name") or ""
            if not confirmed_at and not confirmed_by_name:
                continue
            cleaned_confirmations[department_key] = {
                "confirmed_by_id": status.get("confirmed_by_id"),
                "confirmed_by_name": confirmed_by_name,
                "confirmed_at": confirmed_at,
            }

        locked_at = existing.get("locked_at") or ""
        locked_by_name = existing.get("locked_by_name") or ""
        if cleaned_confirmations and not locked_at:
            first_confirmation = sorted(
                cleaned_confirmations.values(),
                key=lambda item: item.get("confirmed_at") or "",
            )[0]
            locked_at = first_confirmation.get("confirmed_at") or ""
            locked_by_name = first_confirmation.get("confirmed_by_name") or ""

        normalized.append(
            {
                "stt": idx + 1,
                "owner": default["owner"],
                "category": default["category"],
                "detail": str(existing.get("detail", default["detail"]) or ""),
                "note": str(existing.get("note", default["note"]) or ""),
                "department_keys": list(department_keys),
                "confirmations": cleaned_confirmations,
                "is_locked": bool(cleaned_confirmations),
                "locked_at": locked_at,
                "locked_by_name": locked_by_name,
            }
        )
    return normalized


def _write_log(*, plan, row, action, actor=None, department_key="", detail_before="", detail_after="", note_before="", note_after="", extra_data=None):
    ImplementationPlanLog.objects.create(
        plan=plan,
        row_stt=row.get("stt"),
        row_owner=row.get("owner", ""),
        row_category=row.get("category", ""),
        action=action,
        department_key=department_key or "",
        department_label=get_department_label(department_key) if department_key else "",
        actor=actor if getattr(actor, "is_authenticated", False) else None,
        actor_name=_actor_display_name(actor),
        detail_before=detail_before or "",
        detail_after=detail_after or "",
        note_before=note_before or "",
        note_after=note_after or "",
        extra_data=extra_data or {},
    )


def get_or_create_plan(contract):
    plan, _ = ImplementationPlan.objects.get_or_create(contract=contract)
    normalized = normalize_rows(plan.rows_json or [], contract)
    if plan.rows_json != normalized:
        plan.rows_json = normalized
        plan.save(update_fields=["rows_json", "updated_at"])
    return plan


def update_plan_rows_from_post(plan, post_data, contract, actor=None):
    defaults = build_default_rows(contract)
    current_rows = normalize_rows(plan.rows_json or [], contract)

    details = post_data.getlist("row_detail[]")
    notes = post_data.getlist("row_note[]")

    rows = []
    changed_rows = []
    for idx, default in enumerate(defaults):
        existing = current_rows[idx] if idx < len(current_rows) else default
        is_locked = bool(existing.get("is_locked"))

        if is_locked:
            detail = existing.get("detail", default["detail"])
            note = existing.get("note", default["note"])
        else:
            detail = details[idx].strip() if idx < len(details) and details[idx] is not None else existing.get("detail", default["detail"])
            note = notes[idx].strip() if idx < len(notes) and notes[idx] is not None else existing.get("note", default["note"])

        row = {
            "stt": idx + 1,
            "owner": default["owner"],
            "category": default["category"],
            "detail": detail,
            "note": note,
            "department_keys": existing.get("department_keys") or default["department_keys"],
            "confirmations": existing.get("confirmations") or {},
            "is_locked": is_locked,
            "locked_at": existing.get("locked_at") or "",
            "locked_by_name": existing.get("locked_by_name") or "",
        }
        rows.append(row)

        if (existing.get("detail", "") != detail) or (existing.get("note", "") != note):
            changed_rows.append((existing, row))

    with transaction.atomic():
        plan.rows_json = rows
        plan.save(update_fields=["rows_json", "updated_at"])

        for before, after in changed_rows:
            _write_log(
                plan=plan,
                row=after,
                action=ImplementationPlanLog.ACTION_EDIT,
                actor=actor,
                detail_before=before.get("detail", ""),
                detail_after=after.get("detail", ""),
                note_before=before.get("note", ""),
                note_after=after.get("note", ""),
            )

    return plan


def confirm_plan_row(plan, contract, *, row_stt, department_key, actor):
    rows = normalize_rows(plan.rows_json or [], contract)
    row_index = int(row_stt) - 1
    if row_index < 0 or row_index >= len(rows):
        raise ValueError("Dòng xác nhận không hợp lệ.")

    row = rows[row_index]
    if department_key not in (row.get("department_keys") or []):
        raise ValueError("Phòng ban không thuộc nhiệm vụ của dòng này.")

    current_status = (row.get("confirmations") or {}).get(department_key) or {}
    if current_status.get("confirmed_at"):
        return False, row

    confirmed_at = timezone.now().isoformat()
    confirmations = dict(row.get("confirmations") or {})
    confirmations[department_key] = {
        "confirmed_by_id": getattr(actor, "id", None),
        "confirmed_by_name": _actor_display_name(actor),
        "confirmed_at": confirmed_at,
    }
    row["confirmations"] = confirmations
    row["is_locked"] = True
    if not row.get("locked_at"):
        row["locked_at"] = confirmed_at
        row["locked_by_name"] = _actor_display_name(actor)

    with transaction.atomic():
        plan.rows_json = rows
        plan.save(update_fields=["rows_json", "updated_at"])
        _write_log(
            plan=plan,
            row=row,
            action=ImplementationPlanLog.ACTION_CONFIRM,
            actor=actor,
            department_key=department_key,
            detail_after=row.get("detail", ""),
            note_after=row.get("note", ""),
            extra_data={
                "confirmations": confirmations,
                "locked_at": row.get("locked_at", ""),
                "locked_by_name": row.get("locked_by_name", ""),
            },
        )

    return True, row


def unlock_plan_row(plan, contract, *, row_stt, actor):
    rows = normalize_rows(plan.rows_json or [], contract)
    row_index = int(row_stt) - 1
    if row_index < 0 or row_index >= len(rows):
        raise ValueError("Dòng mở khóa không hợp lệ.")

    row = rows[row_index]
    existing_confirmations = dict(row.get("confirmations") or {})
    if not existing_confirmations:
        return False, row

    row["confirmations"] = {}
    row["is_locked"] = False
    row["locked_at"] = ""
    row["locked_by_name"] = ""

    with transaction.atomic():
        plan.rows_json = rows
        plan.save(update_fields=["rows_json", "updated_at"])
        _write_log(
            plan=plan,
            row=row,
            action=ImplementationPlanLog.ACTION_UNLOCK,
            actor=actor,
            detail_after=row.get("detail", ""),
            note_after=row.get("note", ""),
            extra_data={
                "removed_confirmations": existing_confirmations,
            },
        )

    return True, row


def build_plan_rows_for_display(plan, contract, user, *, can_edit=False, is_executive=False):
    rows = normalize_rows(plan.rows_json or [], contract)
    user_department_keys = set(get_user_department_keys(user))
    is_sale_owner = bool(can_edit)
    is_superuser = bool(getattr(user, "is_superuser", False))

    display_rows = []
    for row in rows:
        row_departments = row.get("department_keys") or []
        confirmations = row.get("confirmations") or {}

        department_items = []
        for department_key in row_departments:
            status = confirmations.get(department_key) or {}
            is_confirmed = bool(status.get("confirmed_at"))
            can_click = (
                department_key in user_department_keys
                and not is_sale_owner
                and not is_executive
                and not is_superuser
            )
            visible = is_sale_owner or is_executive or is_superuser or can_click
            if not visible:
                continue

            department_items.append(
                {
                    "key": department_key,
                    "label": get_department_label(department_key),
                    "is_confirmed": is_confirmed,
                    "confirmed_at": status.get("confirmed_at", ""),
                    "confirmed_by_name": status.get("confirmed_by_name", ""),
                    "can_click": can_click and not is_confirmed,
                }
            )

        display_rows.append(
            {
                **row,
                "is_locked": bool(row.get("confirmations")),
                "is_editable": bool(can_edit and not row.get("confirmations")),
                "department_items": department_items,
                "confirmed_count": sum(1 for item in confirmations.values() if item.get("confirmed_at")),
            }
        )
    return display_rows


def build_plan_log_entries(logs):
    entries = []
    for log in logs:
        description = ""
        if log.action == ImplementationPlanLog.ACTION_CONFIRM:
            description = f"{log.actor_name} đã xác nhận phòng ban {log.department_label or '-'}"
        elif log.action == ImplementationPlanLog.ACTION_UNLOCK:
            removed = log.extra_data.get("removed_confirmations") or {}
            removed_labels = [get_department_label(key) for key in removed.keys()]
            labels_text = ", ".join(removed_labels) if removed_labels else "toàn bộ xác nhận"
            description = f"{log.actor_name} đã gỡ xác nhận / mở khóa ({labels_text})"
        elif log.action == ImplementationPlanLog.ACTION_EDIT:
            changed_parts = []
            if (log.detail_before or "") != (log.detail_after or ""):
                changed_parts.append("Chi tiết")
            if (log.note_before or "") != (log.note_after or ""):
                changed_parts.append("Ghi chú")
            description = f"{log.actor_name} đã cập nhật: {', '.join(changed_parts) if changed_parts else 'Nội dung'}"

        entries.append(
            {
                "id": log.id,
                "created_at": log.created_at,
                "row_stt": log.row_stt,
                "row_owner": log.row_owner,
                "row_category": log.row_category,
                "action": log.action,
                "action_label": log.get_action_display(),
                "department_label": log.department_label,
                "actor_name": log.actor_name,
                "description": description,
                "detail_before": log.detail_before,
                "detail_after": log.detail_after,
                "note_before": log.note_before,
                "note_after": log.note_after,
            }
        )
    return entries


def build_package_rows(contract):
    rows = []
    details = contract.service_lines.order_by("display_order", "id")
    for idx, item in enumerate(details, start=1):
        rows.append(
            {
                "stt": idx,
                "group_name": item.group_name or "Khác",
                "item_name": item.item_name or "",
                "description": item.description or "",
                "male": bool(item.for_male),
                "female_single": bool(item.for_female_single),
                "female_family": bool(item.for_female_family),
            }
        )
    return rows


def build_sheet_year(contract):
    if contract.start_date:
        return contract.start_date.year
    if getattr(contract, "created_at", None):
        return contract.created_at.year
    return 2026


def _slugify_ascii(value):
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "ke-hoach-trien-khai"


def _apply_table_border(ws, start_row, end_row, start_col, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def export_plan_excel(contract, plan):
    year = build_sheet_year(contract)
    profile = getattr(contract, "corporate_profile", None)

    company_name = ""
    if profile and getattr(profile, "company_name_snapshot", None):
        company_name = profile.company_name_snapshot
    elif contract.company:
        company_name = contract.company.name

    wb = Workbook()
    ws1 = wb.active
    ws1.title = f"TRIỂN KHAI {year}"
    ws2 = wb.create_sheet(title=f"GÓI KHÁM {year}")

    rows = normalize_rows(plan.rows_json or [], contract)
    package_rows = build_package_rows(contract)

    ws1.merge_cells("B1:E1")
    ws1["B1"] = f"TRIỂN KHAI KHÁM SỨC KHỎE NĂM {year}"
    ws1["B1"].font = Font(bold=True, size=14, color="0F172A")
    ws1["B1"].fill = TITLE_FILL
    ws1["B1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1["B2"] = "Công ty"
    ws1["C2"] = company_name or "-"
    ws1["D2"] = "Số hợp đồng"
    ws1["E2"] = contract.contract_number or "-"

    ws1.append([None, "STT", "Phụ trách", "Danh mục", "Chi tiết", "Ghi chú / Giải thích"])
    header_row = ws1.max_row
    _style_header_row(ws1, header_row)

    for row in rows:
        ws1.append([
            None,
            row.get("stt"),
            row.get("owner"),
            row.get("category"),
            row.get("detail"),
            row.get("note"),
        ])

    _apply_table_border(ws1, header_row, ws1.max_row, 2, 6)
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 34
    ws1.column_dimensions["E"].width = 55
    ws1.column_dimensions["F"].width = 38

    ws2.merge_cells("B1:E1")
    ws2["B1"] = f"GÓI KHÁM {year} - {company_name or '-'}"
    ws2["B1"].font = Font(bold=True, size=14, color="0F172A")
    ws2["B1"].fill = TITLE_FILL
    ws2["B1"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.append([None, "STT", "Danh mục khám", "Nam", "Nữ độc thân", "Nữ có gia đình"])
    header_row = ws2.max_row
    _style_header_row(ws2, header_row)

    current_group = None
    for item in package_rows:
        if item["group_name"] != current_group:
            current_group = item["group_name"]
            ws2.append([None, current_group, None, None, None, None])
            group_row_idx = ws2.max_row
            for cell in ws2[group_row_idx][1:6]:
                cell.fill = GROUP_FILL
                cell.font = Font(bold=True, color="0F4C81")
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws2.merge_cells(start_row=group_row_idx, start_column=2, end_row=group_row_idx, end_column=6)

        ws2.append([
            None,
            item.get("stt"),
            item.get("item_name"),
            "✓" if item.get("male") else "",
            "✓" if item.get("female_single") else "",
            "✓" if item.get("female_family") else "",
        ])

    _apply_table_border(ws2, header_row, ws2.max_row, 2, 6)
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 54
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 18

    for ws in (ws1, ws2):
        for row in ws.iter_rows():
            for cell in row:
                if cell.coordinate in {"B1", "B2", "C2", "D2", "E2"}:
                    continue
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{_slugify_ascii(company_name)}-ke-hoach-trien-khai-{year}.xlsx"
    return output, filename
